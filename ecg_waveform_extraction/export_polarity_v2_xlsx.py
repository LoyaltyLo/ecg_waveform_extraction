"""Generate final polarity detection Excel report from _polarity_v2.json.

7 sheets: Summary, Per-Record, P-Axis, QRS-Axis, P-Wave, Lead-I, aVR, Voting, Conflicting
"""
import sys
sys.path.insert(0, 'c:/LoyaltyLo/PythonProjects/ECG_engineering')

import os, json, numpy as np
from collections import Counter

OUT_DIR = 'c:/LoyaltyLo/PythonProjects/ECG_engineering/ecg_waveform_extraction/output_rala_full'
JSON_PATH = os.path.join(OUT_DIR, '_polarity_v2.json')
XLSX_PATH = os.path.join(OUT_DIR, 'polarity_detection_v2.xlsx')

with open(JSON_PATH) as f:
    all_data = json.load(f)

total = len(all_data)

# ---- xlsx ----
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.series import DataPoint

HDR_FONT = Font(name='Consolas', bold=True, color='FFFFFF', size=10)
HDR_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
SEC_FONT = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=12)
SEC_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
GRN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
YEL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
BLU = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
GRY = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
BORDER = Border(left=Side('thin','D0D0D0'), right=Side('thin','D0D0D0'),
                top=Side('thin','D0D0D0'), bottom=Side('thin','D0D0D0'))

def hdr(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = HDR_ALIGN; cell.border = BORDER

def sec(ws, row, text, ncols=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row=row, column=1, value=text).font = SEC_FONT
    for c in range(1, ncols+1): ws.cell(row=row, column=c).fill = SEC_FILL
    ws.row_dimensions[row].height = 26

def cell(ws, row, col, val, font=None, align=None, fill=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font or Font(name='Consolas', size=10)
    c.alignment = align or Alignment(horizontal='center', vertical='center')
    c.border = BORDER
    if fill: c.fill = fill
    return c

def auto_w(ws, mn=9, mx=55):
    for col_cells in ws.columns:
        cl = get_column_letter(col_cells[0].column)
        m = mn
        for c in col_cells:
            if c.value:
                for line in str(c.value).split('\n'):
                    m = max(m, min(len(line)+2, mx))
        ws.column_dimensions[cl].width = m

wb = Workbook()

# ============ Sheet 1: Summary ============
ws = wb.active
ws.title = 'Summary'
ws.sheet_properties.tabColor = '2F5496'
ws.merge_cells('A1:J1')
ws['A1'].value = 'RA-LA Reversal aECG — 5-Method Polarity Detection Report'
ws['A1'].font = Font(name='Microsoft YaHei', bold=True, size=16, color='1F4E79')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

n_n = sum(1 for d in all_data if d['consensus'] == 'normal')
n_r = sum(1 for d in all_data if d['consensus'] == 'reversed')
n_u = sum(1 for d in all_data if d['consensus'] == 'uncertain')

# Method comparison table
row = 3
sec(ws, row, '🔍 5种检测方法对比', 8)
row = 4
for j, h in enumerate(['方法','原理','正常','反接','权重','检测反接数','反接占比','可靠性']):
    ws.cell(row=row, column=j+1, value=h)
hdr(ws, row, 8)

methods_summary = [
    ['① Lead I P/QRS', 'Lead I 导联P波+QRS极性', 'P↑ QRS↑', 'P↓ QRS↓(同时)', '×1.3',
     sum(1 for d in all_data if d['methods'].get('lead_I',{}).get('polarity')=='reversed'),
     '—', '⭐⭐⭐ 最可靠单导联'],
    ['② aVR 导联', 'aVR P波极性 (正常永为负)', 'P↓ QRS↓', 'P↑→病征性反接', '×1.3',
     sum(1 for d in all_data if d['methods'].get('avr_lead',{}).get('polarity') in ('reversed',)),
     '—', '⭐⭐⭐ P↑高度特异'],
    ['③ P波电轴', '额面P波电轴角度', '0°~75°', '>100° 或 <-30°', '×0.9',
     sum(1 for d in all_data if d['methods'].get('p_axis',{}).get('polarity')=='reversed'),
     '—', '⭐⭐ 机器测量可能有误差'],
    ['④ QRS电轴', '额面QRS电轴角度', '-30°~+90°', '>120° 极度右偏', '×0.7',
     sum(1 for d in all_data if d['methods'].get('qrs_axis',{}).get('polarity')=='reversed'),
     '—', '⭐⭐ 受传导阻滞干扰'],
    ['⑤ P波极性(HSMM)', 'Lead II P波HSMM分割积分', '正向 (↑)', '负向 (↓)', '×1.0',
     sum(1 for d in all_data if d['methods'].get('p_wave_hsmm',{}).get('polarity')=='reversed'),
     '—', '⭐⭐ 需窦性心律'],
]
for i, md in enumerate(methods_summary):
    for j, val in enumerate(md):
        cell(ws, row+1+i, j+1, val,
             font=Font(name='Microsoft YaHei', size=10),
             align=Alignment(horizontal='center' if j!=1 else 'left', wrap_text=True))
    ws.row_dimensions[row+1+i].height = 32

# Consensus section
row = 12
sec(ws, row, '📋 综合投票结果', 8)
row = 13
for j, h in enumerate(['判定','记录数','占比','说明']):
    ws.cell(row=row, column=j+1, value=h)
hdr(ws, row, 4)

for i, (label, n, fill, desc) in enumerate([
    ('极性正常', n_n, GRN, '5方法加权投票中"正常"票占多数'),
    ('极性反接 (RA-LA Reversal)', n_r, RED, '5方法加权投票中"反接"票占多数，需重新采集'),
    ('不确定', n_u, YEL, '票数不足或方法结果矛盾，需人工复核'),
]):
    for j, val in enumerate([label, n, f'{n/max(total,1)*100:.1f}%', desc]):
        cell(ws, row+1+i, j+1, val,
             font=Font(name='Microsoft YaHei', size=11, bold=(j==0)),
             align=Alignment(horizontal='center' if j<3 else 'left'),
             fill=fill)

# Pie chart
row = 19
ws.merge_cells(f'A{row}:C{row}')
ws.cell(row=row, column=1, value='综合判定分布').font = Font(name='Microsoft YaHei', bold=True, size=11)
ws.cell(row=row+1, column=1, value='分类'); ws.cell(row=row+1, column=2, value='数量')
ws.cell(row=row+2, column=1, value='极性正常'); ws.cell(row=row+2, column=2, value=n_n)
ws.cell(row=row+3, column=1, value='极性反接'); ws.cell(row=row+3, column=2, value=n_r)
ws.cell(row=row+4, column=1, value='不确定'); ws.cell(row=row+4, column=2, value=n_u)

pie = PieChart(); pie.title='RA-LA 极性综合判定'; pie.style=10
pie.add_data(Reference(ws, min_col=2, min_row=row+1, max_row=row+4), titles_from_data=True)
pie.set_categories(Reference(ws, min_col=1, min_row=row+2, max_row=row+4))
pie.width=18; pie.height=12
for i, color in enumerate(['C6EFCE','FFC7CE','FFEB9C']):
    pt=DataPoint(idx=i); pt.graphicalProperties.solidFill=color; pie.series[0].data_points.append(pt)
ws.add_chart(pie, f'E{row}')
ws.row_dimensions[row].height = 280

# Key stats
row = 26
sec(ws, row, '📈 关键统计', 8)
row = 27
axes_p = [d.get('p_axis') for d in all_data if d.get('p_axis') is not None]
axes_q = [d.get('qrs_axis') for d in all_data if d.get('qrs_axis') is not None]
hrs = [d.get('hr') for d in all_data if d.get('hr') is not None]
prs = [d.get('pr') for d in all_data if d.get('pr') is not None]
qrss = [d.get('qrs') for d in all_data if d.get('qrs') is not None]

stats = [
    ['P轴均值', f'{np.mean(axes_p):.0f}°', 'QRS轴均值', f'{np.mean(axes_q):.0f}°', '', ''],
    ['正常组P轴', f'{np.mean([a for a in axes_p if 0<=a<=75]):.0f}°', '反接组P轴', f'{np.mean([a for a in axes_p if a>100 or a<-30]):.0f}°', '', ''],
    ['HR均值', f'{np.mean(hrs):.1f} bpm', 'PR均值', f'{np.mean(prs):.1f} ms', 'QRS均值', f'{np.mean(qrss):.1f} ms'],
    ['总记录', total, 'Lead I权重', '×1.3', 'aVR权重', '×1.3'],
]
for j, h in enumerate(['指标','值','指标','值','指标','值']):
    ws.cell(row=row, column=j+1, value=h)
hdr(ws, row, 6)
for i, st in enumerate(stats):
    for j, val in enumerate(st):
        cell(ws, row+1+i, j+1, val, font=Font(name='Consolas',size=10), fill=GRY if j%2==0 else None)
auto_w(ws, 14, 55)
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 28


# ============ Sheet 2: Per-Record ============
ws2 = wb.create_sheet('Per-Record')
ws2.sheet_properties.tabColor = '4472C4'
hdrs2 = ['#','Record','Consensus','Confidence','P-Axis(°)','QRS-Axis(°)',
         'P-Axis Pol','QRS-Axis Pol','Lead I Pol','aVR Pol','P-Wave Pol',
         'Lead I P_net','Lead I QRS_net','aVR P_net','aVR QRS_net',
         'P-Wave +','P-Wave -','P-Wave net','HR','PR','QRS','QT','Interpretation']
for j, h in enumerate(hdrs2):
    ws2.cell(row=1, column=j+1, value=h)
hdr(ws2, 1, len(hdrs2))
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f'A1:{get_column_letter(len(hdrs2))}{total+1}'

for i, d in enumerate(all_data):
    r = i + 2
    m = d.get('methods', {})
    li = m.get('lead_I', {})
    av = m.get('avr_lead', {})
    pw = m.get('p_wave_hsmm', {})
    pa = m.get('p_axis', {})
    qa = m.get('qrs_axis', {})

    vals = [
        i+1, d['record'], d.get('consensus','?'), d.get('confidence',0),
        d.get('p_axis'), d.get('qrs_axis'),
        pa.get('polarity',''), qa.get('polarity',''),
        li.get('polarity',''), av.get('polarity',''), pw.get('polarity',''),
        li.get('p_net'), li.get('qrs_net'), av.get('p_net'), av.get('qrs_net'),
        pw.get('n_pos'), pw.get('n_neg'), pw.get('mean_area'),
        d.get('hr'), d.get('pr'), d.get('qrs'), d.get('qt'),
        (d.get('interpretation','') or '')[:180],
    ]
    for j, val in enumerate(vals):
        cell(ws2, r, j+1, val, align=Alignment(horizontal='left',wrap_text=True) if j==len(vals)-1 else None)

    # Color by consensus
    cons = d.get('consensus','')
    if cons == 'reversed':
        for j in [3]: ws2.cell(row=r, column=j).fill = RED
    elif cons == 'normal':
        for j in [3]: ws2.cell(row=r, column=j).fill = GRN
    elif cons == 'uncertain':
        for j in [3]: ws2.cell(row=r, column=j).fill = YEL
    # Color individual method cells
    if li.get('polarity') == 'reversed':
        ws2.cell(row=r, column=9).fill = RED
    if av.get('polarity') == 'reversed':
        ws2.cell(row=r, column=10).fill = RED
    if pw.get('polarity') == 'reversed':
        ws2.cell(row=r, column=11).fill = RED

auto_w(ws2, 8, 45)
ws2.column_dimensions[get_column_letter(len(hdrs2))].width = 55


# ============ Sheets 3-7: Per-Method Detail ============
method_sheets = [
    ('P-Axis', 'p_axis', '5B9BD5',
     ['Record','P-Axis(°)','Polarity','Confidence','Detail','Consensus','Match?','HR']),
    ('QRS-Axis', 'qrs_axis', 'ED7D31',
     ['Record','QRS-Axis(°)','Polarity','Confidence','Detail','Consensus','Match?','HR']),
    ('P-Wave HSMM', 'p_wave_hsmm', '70AD47',
     ['Record','Polarity','N+','N-','Net Area','Confidence','Detail','Consensus','Match?']),
    ('Lead I', 'lead_I', 'FFC000',
     ['Record','Polarity','P Inv','P Net','QRS Inv','QRS Net','Confidence','Detail','Consensus','Match?']),
    ('aVR Lead', 'avr_lead', '7030A0',
     ['Record','Polarity','P Pos','P Net','QRS Pos','QRS Net','Confidence','Detail','Consensus','Match?']),
]

for sheet_name, method_key, tab_color, headers in method_sheets:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color
    for j, h in enumerate(headers):
        ws.cell(row=1, column=j+1, value=h)
    hdr(ws, 1, len(headers))
    ws.freeze_panes = 'A2'

    for i, d in enumerate(all_data):
        r = i + 2
        m = d.get('methods', {}).get(method_key, {})
        consensus = d.get('consensus', '')
        pol = m.get('polarity', '')
        match = '✓' if pol and consensus in ('normal','reversed') and pol in (consensus, f'likely_{consensus}') else ('✗' if pol else '—')

        if method_key == 'p_axis':
            vals = [d['record'], d.get('p_axis'), pol, m.get('conf'),
                    m.get('detail',''), consensus, match, d.get('hr')]
            if m.get('polarity') == 'reversed':
                ws.cell(row=r, column=2).fill = RED
                ws.cell(row=r, column=3).fill = RED
        elif method_key == 'qrs_axis':
            vals = [d['record'], d.get('qrs_axis'), pol, m.get('conf'),
                    m.get('detail',''), consensus, match, d.get('hr')]
            if m.get('polarity') == 'reversed':
                ws.cell(row=r, column=2).fill = RED
                ws.cell(row=r, column=3).fill = RED
        elif method_key == 'p_wave_hsmm':
            vals = [d['record'], pol, m.get('n_pos'), m.get('n_neg'),
                    m.get('mean_area'), m.get('conf'), m.get('detail',''),
                    consensus, match]
        elif method_key == 'lead_I':
            vals = [d['record'], pol, 'Yes' if m.get('p_inverted') else ('No' if m.get('p_inverted')==False else ''),
                    m.get('p_net'), 'Yes' if m.get('qrs_inverted') else ('No' if m.get('qrs_inverted')==False else ''),
                    m.get('qrs_net'), m.get('conf'), m.get('detail',''), consensus, match]
        elif method_key == 'avr_lead':
            vals = [d['record'], pol, 'Yes' if m.get('p_positive') else ('No' if m.get('p_positive')==False else ''),
                    m.get('p_net'), 'Yes' if m.get('qrs_positive') else ('No' if m.get('qrs_positive')==False else ''),
                    m.get('qrs_net'), m.get('conf'), m.get('detail',''), consensus, match]

        for j, val in enumerate(vals):
            cell(ws, r, j+1, val)

        if pol == 'reversed':
            ws.cell(row=r, column=2).fill = RED
        elif pol == 'normal':
            ws.cell(row=r, column=2).fill = GRN
        if match == '✗':
            cell(ws, r, len(vals), match, fill=RED, font=Font(name='Consolas',bold=True,size=12,color='C00000'))
        elif match == '✓':
            cell(ws, r, len(vals), match, font=Font(name='Consolas',bold=True,size=12,color='006100'))

    auto_w(ws, 8, 55)


# ============ Sheet 8: Combined Voting ============
ws8 = wb.create_sheet('Voting')
ws8.sheet_properties.tabColor = '00B050'
hdrs8 = ['Record','P-Axis Vote','QRS-Axis Vote','Lead I Vote','aVR Vote','P-Wave Vote',
         'Votes N','Votes R','Consensus','Confidence','P-Axis(°)','QRS-Axis(°)','Interpretation']
for j, h in enumerate(hdrs8):
    ws8.cell(row=1, column=j+1, value=h)
hdr(ws8, 1, len(hdrs8))
ws8.freeze_panes = 'A2'

for i, d in enumerate(all_data):
    r = i + 2
    m = d.get('methods', {})
    votes = []
    for mk in ['p_axis','qrs_axis','lead_I','avr_lead','p_wave_hsmm']:
        pol = m.get(mk, {}).get('polarity', '')
        if pol == 'normal': v = 'N'
        elif pol in ('reversed','likely_reversed'): v = 'R'
        elif pol == 'borderline': v = '~'
        else: v = '-'
        votes.append(v)

    vals = [d['record']] + votes + [
        round(d['votes']['normal'],2), round(d['votes']['reversed'],2),
        d.get('consensus',''), d.get('confidence',0),
        d.get('p_axis'), d.get('qrs_axis'),
        (d.get('interpretation','') or '')[:150],
    ]
    for j, val in enumerate(vals):
        cell(ws8, r, j+1, val, align=Alignment(horizontal='left',wrap_text=True) if j==len(vals)-1 else None)

    cons = d.get('consensus','')
    if cons == 'reversed':
        ws8.cell(row=r, column=8).fill = RED
    elif cons == 'normal':
        ws8.cell(row=r, column=8).fill = GRN

    # Color individual votes
    for j, v in enumerate(votes):
        if v == 'R': ws8.cell(row=r, column=2+j).fill = RED
        elif v == 'N': ws8.cell(row=r, column=2+j).fill = GRN

auto_w(ws8, 8, 50)
ws8.column_dimensions[get_column_letter(len(hdrs8))].width = 50


# ============ Sheet 9: Conflicting ============
ws9 = wb.create_sheet('Conflicting')
ws9.sheet_properties.tabColor = 'FF0000'
hdrs9 = ['Record','Consensus','Conflict Description',
         'P-Axis','QRS-Axis','Lead I','aVR','P-Wave',
         'P-Axis(°)','QRS-Axis(°)','Lead I P_net','aVR P_net','HR','Interpretation']
for j, h in enumerate(hdrs9):
    ws9.cell(row=1, column=j+1, value=h)
hdr(ws9, 1, len(hdrs9))
ws9.freeze_panes = 'A2'

cr = 2
for d in all_data:
    m = d.get('methods', {})
    pols = {}
    for mk in ['p_axis','qrs_axis','lead_I','avr_lead','p_wave_hsmm']:
        p = m.get(mk, {}).get('polarity', '')
        if p in ('normal',): pols[mk] = 'N'
        elif p in ('reversed','likely_reversed'): pols[mk] = 'R'
        elif p: pols[mk] = '~'

    has_N = 'N' in pols.values()
    has_R = 'R' in pols.values()
    if not (has_N and has_R):
        continue  # skip unanimous records

    li = m.get('lead_I', {})
    av = m.get('avr_lead', {})

    conflict_desc = ' vs '.join(f'{k}={v}' for k,v in pols.items() if v in ('N','R'))

    vals = [
        d['record'], d.get('consensus',''), conflict_desc,
        pols.get('p_axis','-'), pols.get('qrs_axis','-'),
        pols.get('lead_I','-'), pols.get('avr_lead','-'), pols.get('p_wave_hsmm','-'),
        d.get('p_axis'), d.get('qrs_axis'),
        li.get('p_net'), av.get('p_net'),
        d.get('hr'), (d.get('interpretation','') or '')[:150],
    ]
    for j, val in enumerate(vals):
        cell(ws9, cr, j+1, val, fill=YEL, align=Alignment(horizontal='left',wrap_text=True) if j==len(vals)-1 else None)
    cr += 1

if cr == 2:
    ws9.merge_cells('A2:N2')
    ws9.cell(row=2, column=1, value='无矛盾记录 — 全部5种方法结果一致').font = Font(name='Microsoft YaHei',bold=True,size=12,color='006100')
    ws9['A2'].fill = GRN

auto_w(ws9, 8, 55)
ws9.column_dimensions[get_column_letter(len(hdrs9))].width = 50
ws9.column_dimensions['C'].width = 35

# ---- Freeze & save ----
for ws in wb.worksheets[1:]:
    try: ws.freeze_panes = 'A2'
    except: pass

wb.save(XLSX_PATH)
size_mb = os.path.getsize(XLSX_PATH) / (1024*1024)
print(f"Saved: {XLSX_PATH} ({size_mb:.1f} MB)")
print(f"  Sheets: Summary, Per-Record, P-Axis, QRS-Axis, P-Wave HSMM, Lead I, aVR Lead, Voting, Conflicting")
print(f"  Records: {total}")
print(f"  Normal: {n_n} ({n_n/max(total,1)*100:.1f}%) | Reversed: {n_r} ({n_r/max(total,1)*100:.1f}%) | Uncertain: {n_u}")
