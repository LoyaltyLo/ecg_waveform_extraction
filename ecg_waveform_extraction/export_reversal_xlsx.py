#!/usr/bin/env python3
"""Export limb lead reversal detection results to detailed Excel (.xlsx).

Usage:
    python -m ecg_waveform_extraction.export_reversal_xlsx
    python -m ecg_waveform_extraction.export_reversal_xlsx --n 50
    python -m ecg_waveform_extraction.export_reversal_xlsx --from-cache
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import os, json, time, gc, argparse
from collections import Counter
import numpy as np

from ecg_waveform_extraction.limb_lead_processor import (
    LimbLeadProcessor, LIMB_LEADS,
)
from ecg_waveform_extraction.limb_lead_reversal import (
    LimbLeadReversalDetector, REVERSAL_TYPES, REVERSAL_NAMES,
)
from ecg_waveform_extraction.utils.aecg_parser import parse_aecg

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
AECG_DIR = 'C:/LoyaltyLo/datasets/RA-LA_Reversal/aECG'
CACHE_DIR = str(Path(__file__).resolve().parent / 'output_rala_full/_limb_leads')
OUT_DIR = str(Path(__file__).resolve().parent / 'output_rala_full/_reversal')
MAX_SAMPLES = 4000

# ---------------------------------------------------------------------------
# Style definitions
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name='Consolas', size=10, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

DATA_FONT = Font(name='Consolas', size=9)
DATA_ALIGN = Alignment(horizontal='center', vertical='center')
TEXT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Verdict colors
FILL_REVERSED = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')  # light red
FILL_NORMAL = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')    # light green
FILL_UNCERTAIN = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')  # light yellow
FILL_HIGH_CONF = PatternFill(start_color='B71C1C', end_color='B71C1C', fill_type='solid')  # dark red for high-conf reversal

FONT_WHITE = Font(name='Consolas', size=9, bold=True, color='FFFFFF')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

# ---------------------------------------------------------------------------
# Extract patient name from aECG
# ---------------------------------------------------------------------------
def get_patient_name(filepath: str) -> str:
    """Extract patient name from aECG XML file."""
    import re
    try:
        with open(filepath, 'rb') as f:
            content = f.read()
        # Patient name is in the first <name> tag within <patient> block
        m = re.search(rb'<patient[^>]*>.*?<name[^>]*>([^<]+)</name>', content, re.DOTALL)
        if m:
            return m.group(1).decode('utf-8', errors='replace')
    except Exception:
        pass
    return ''


# ---------------------------------------------------------------------------
# Load cached LimbLeadResult
# ---------------------------------------------------------------------------
def load_cached_result(rec_name: str):
    """Reconstruct LimbLeadResult from cache."""
    cache_path = os.path.join(CACHE_DIR, rec_name, 'summary.json')
    if not os.path.exists(cache_path):
        return None

    with open(cache_path, encoding='utf-8') as f:
        cache_data = json.load(f)

    from ecg_waveform_extraction.limb_lead_processor import LeadResult, LimbLeadResult

    leads = {}
    for ln, ld in cache_data['leads'].items():
        if ld is None:
            leads[ln] = None
        else:
            leads[ln] = LeadResult(
                lead_name=ld['lead_name'], n_beats=ld['n_beats'],
                polarity_counts=ld['polarity_counts'],
                p_polarity_counts=ld['p_polarity_counts'],
                t_polarity_counts=ld.get('t_polarity_counts', {}),
                mean_qrs_dur_ms=ld['mean_qrs_dur_ms'],
                mean_rs_ratio=ld['mean_rs_ratio'],
                mean_qrs_net=ld['mean_qrs_net'],
                mean_p_net=ld['mean_p_net'],
                mean_t_net=ld.get('mean_t_net', 0.0),
                mean_r_amplitude=ld.get('mean_r_amplitude', 0),
                mean_s_amplitude=ld.get('mean_s_amplitude', 0),
            )

    return LimbLeadResult(
        record=cache_data['record'], fs=cache_data['fs'],
        n_samples=cache_data['n_samples'], leads=leads,
        measurements=cache_data.get('measurements', {}),
        interpretation=cache_data.get('interpretation', ''),
        n_total_beats=cache_data['n_total_beats'],
    )


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------
def export_to_xlsx(results: list, detector, output_path: str,
                   patient_names: dict | None = None):
    """Generate multi-sheet detailed Excel report.

    Sheets:
      1. Overview      — All records, verdict, confidence, key metrics
      2. Criteria      — Per-record, per-type, per-criterion details
      3. Lead Polarity — Per-lead QRS and P-wave dominant polarity
      4. RA-LA Detail  — RA-LA-specific deep dive
    """
    wb = openpyxl.Workbook()

    # ==================================================================
    # Sheet 1: Overview
    # ==================================================================
    ws1 = wb.active
    ws1.title = 'Overview'

    # Headers
    headers = [
        'Record', 'Patient Name', 'Verdict', 'Confidence',
        'P_axis', 'QRS_axis', 'HR',
        'RA-LA Score(N)', 'RA-LA Score(R)', 'RA-LA N_Criteria',
        'Lead I QRS', 'Lead I P', 'Lead I T', 'aVR QRS', 'aVR P',
        'P_axis', 'QRS_axis', 'Interpretation',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Data rows
    for i, r in enumerate(results):
        row = i + 2
        meas = r.measurements

        def _dom(counts):
            if not counts: return 'N/A'
            f = {k:v for k,v in counts.items() if k!='uncertain'}
            return max(f, key=f.get) if f else 'uncertain'

        i_lr = r.lead_polarity_summary.get('I') or {}
        avr_lr = r.lead_polarity_summary.get('AVR') or {}

        i_qrs = _dom(i_lr.get('qrs_counts', {}))
        i_p = _dom(i_lr.get('p_counts', {}))
        i_t = _dom(i_lr.get('t_counts', {}))
        avr_qrs = _dom(avr_lr.get('qrs_counts', {}))
        avr_p = _dom(avr_lr.get('p_counts', {}))

        tr = r.types.get('ra_la')
        ra_la_sn = tr.score_normal if tr else ''
        ra_la_sr = tr.score_reversed if tr else ''
        ra_la_ncrit = tr.n_criteria_triggered if tr else ''

        # Patient name via aECG lookup
        patient_name = patient_names.get(r.record, '') if patient_names else ''

        values = [
            r.record, patient_name, r.reversal_type, r.confidence,
            ra_la_sn, ra_la_sr, ra_la_ncrit,
            i_qrs, i_p, i_t, avr_qrs, avr_p,
            meas.get('P_axis', ''), meas.get('QRS_axis', ''),
            (r.interpretation or '')[:80],
        ]

        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN if not isinstance(val, str) else TEXT_ALIGN
            cell.border = THIN_BORDER

        # Color verdict column
        verdict_cell = ws1.cell(row=row, column=2)
        if r.reversal_type in REVERSAL_TYPES:
            verdict_cell.fill = FILL_REVERSED
            if r.confidence >= 0.8:
                verdict_cell.font = Font(name='Consolas', size=9, bold=True, color='B71C1C')
        elif r.reversal_type == 'normal':
            verdict_cell.fill = FILL_NORMAL
        else:
            verdict_cell.fill = FILL_UNCERTAIN

        # Color confidence
        conf_cell = ws1.cell(row=row, column=3)
        if r.confidence >= 0.8:
            conf_cell.font = Font(name='Consolas', size=9, bold=True, color='2E7D32')

    # Column widths
    col_widths = [14, 10, 10, 10, 8, 9, 6, 12, 12, 10, 8, 8, 8, 8, 8, 8, 8, 40]
    for col, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # Freeze panes
    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(results)+1}'

    # ==================================================================
    # Sheet 2: Criteria Details
    # ==================================================================
    ws2 = wb.create_sheet('Criteria')

    crit_headers = [
        'Record', 'Type', 'Criterion', 'Verdict', 'Confidence',
        'Weight', 'Detail',
    ]
    for col, h in enumerate(crit_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    row = 2
    for r in results:
        for tname in REVERSAL_TYPES:
            tr = r.types.get(tname)
            if tr is None:
                continue
            for c in tr.criteria:
                values = [
                    r.record, tname, c.name, c.verdict, c.confidence,
                    c.weight, c.detail,
                ]
                for col, val in enumerate(values, 1):
                    cell = ws2.cell(row=row, column=col, value=val)
                    cell.font = DATA_FONT
                    cell.alignment = TEXT_ALIGN if isinstance(val, str) else DATA_ALIGN
                    cell.border = THIN_BORDER

                # Color verdict
                vc = ws2.cell(row=row, column=4)
                if c.verdict == 'reversed':
                    vc.fill = FILL_REVERSED
                elif c.verdict == 'normal':
                    vc.fill = FILL_NORMAL
                elif c.verdict == 'borderline':
                    vc.fill = FILL_UNCERTAIN

                row += 1

    crit_widths = [14, 8, 24, 10, 8, 8, 70]
    for col, w in enumerate(crit_widths, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f'A1:{get_column_letter(len(crit_headers))}{row-1}'

    # ==================================================================
    # Sheet 3: Lead Polarity Matrix
    # ==================================================================
    ws3 = wb.create_sheet('Lead Polarity')

    lp_headers = [
        'Record', 'Verdict', 'Confidence',
        'I QRS', 'I P', 'I QRS Counts', 'I P Counts', 'I QRS Net',
        'II QRS', 'II P', 'II QRS Counts', 'II P Counts', 'II QRS Net',
        'III QRS', 'III P', 'III QRS Counts', 'III P Counts', 'III QRS Net',
        'AVR QRS', 'AVR P', 'AVR QRS Counts', 'AVR P Counts', 'AVR QRS Net',
        'AVL QRS', 'AVL P', 'AVL QRS Counts', 'AVL P Counts', 'AVL QRS Net',
        'AVF QRS', 'AVF P', 'AVF QRS Counts', 'AVF P Counts', 'AVF QRS Net',
    ]
    for col, h in enumerate(lp_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for i, r in enumerate(results):
        row = i + 2
        values = [r.record, r.reversal_type, r.confidence]

        for ln in LIMB_LEADS:
            lr = r.lead_polarity_summary.get(ln) or {}
            qrs_dom = lr.get('qrs_dominant', 'N/A')
            p_dom = lr.get('p_dominant', 'N/A')
            qrs_counts = str(lr.get('qrs_counts', {}))
            p_counts = str(lr.get('p_counts', {}))
            qrs_net = lr.get('mean_qrs_net', '')

            values += [qrs_dom, p_dom, qrs_counts, p_counts, qrs_net]

        for col, val in enumerate(values, 1):
            cell = ws3.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.alignment = TEXT_ALIGN if isinstance(val, str) else DATA_ALIGN
            cell.border = THIN_BORDER

        # Color verdict
        vc = ws3.cell(row=row, column=2)
        if r.reversal_type in REVERSAL_TYPES:
            vc.fill = FILL_REVERSED
        elif r.reversal_type == 'normal':
            vc.fill = FILL_NORMAL
        else:
            vc.fill = FILL_UNCERTAIN

        # Color Lead I QRS and aVR QRS based on polarity
        for offset, dom_col in [(3, 4), (21, 22)]:  # Lead I QRS, aVR QRS offset-1
            qrs_cell = ws3.cell(row=row, column=dom_col)
            val_str = str(qrs_cell.value)
            if val_str == 'negative':
                qrs_cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
            elif val_str == 'positive':
                # In aVR, positive is abnormal
                if dom_col == 22:  # aVR QRS
                    qrs_cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                else:
                    qrs_cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')

    lp_widths = [14, 10, 10] + [10, 8, 22, 18, 8] * 6
    for col, w in enumerate(lp_widths, 1):
        ws3.column_dimensions[get_column_letter(col)].width = min(w, 24)
    ws3.freeze_panes = 'C2'

    # ==================================================================
    # Sheet 4: RA-LA Deep Dive
    # ==================================================================
    ws4 = wb.create_sheet('RA-LA Deep Dive')

    rala_headers = [
        'Record', 'RA-LA Verdict', 'RA-LA Score(N)', 'RA-LA Score(R)', 'RA-LA Conf',
        'Lead I P Inverted', 'Conf', 'Detail',
        'aVR P Upright', 'Conf', 'Detail',
        'Lead I QRS+P+T', 'Conf', 'Detail',
        'P-axis', 'Conf', 'Detail',
        'aVR↔aVL Swap', 'Conf', 'Detail',
        'QRS-axis Right', 'Conf', 'Detail',
        'Overall Verdict', 'Overall Conf',
        'P_axis', 'QRS_axis', 'HR',
    ]
    for col, h in enumerate(rala_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for i, r in enumerate(results):
        row = i + 2
        meas = r.measurements

        tr = r.types.get('ra_la')
        if tr is None:
            continue

        values = [
            r.record, tr.verdict, tr.score_normal, tr.score_reversed, tr.confidence,
        ]

        # Each of the 6 RA-LA criteria
        crit_map = {c.name: c for c in tr.criteria}
        for cname in [
            'Lead I P inverted', 'aVR P upright', 'Lead I QRS+P+T',
            'P-axis', 'aVR↔aVL swap', 'QRS-axis right',
        ]:
            c = crit_map.get(cname)
            if c:
                values += [c.verdict, c.confidence, c.detail]
            else:
                values += ['', '', '']

        values += [
            r.reversal_type, r.confidence,
            meas.get('P_axis', ''), meas.get('QRS_axis', ''), meas.get('HR', ''),
        ]

        for col, val in enumerate(values, 1):
            cell = ws4.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.alignment = TEXT_ALIGN if isinstance(val, str) else DATA_ALIGN
            cell.border = THIN_BORDER

            # Color criterion verdict columns
            if col in [6, 9, 12, 15, 18, 21]:  # verdict columns for each criterion
                if val == 'reversed':
                    cell.fill = FILL_REVERSED
                elif val == 'normal':
                    cell.fill = FILL_NORMAL
                elif val == 'borderline':
                    cell.fill = FILL_UNCERTAIN

        # Color overall verdict
        vc = ws4.cell(row=row, column=2)
        if tr.verdict == 'reversed':
            vc.fill = FILL_REVERSED
        elif tr.verdict == 'normal':
            vc.fill = FILL_NORMAL

    rala_widths = [14, 12, 12, 12, 10, 12, 6, 50, 12, 6, 50, 12, 6, 50, 12, 6, 40, 12, 6, 40, 12, 6, 40, 12, 10, 8, 8, 6]
    for col, w in enumerate(rala_widths, 1):
        ws4.column_dimensions[get_column_letter(col)].width = min(w, 50)
    ws4.freeze_panes = 'A2'
    ws4.auto_filter.ref = f'A1:{get_column_letter(len(rala_headers))}{len(results)+1}'

    # ==================================================================
    # Save
    # ==================================================================
    wb.save(output_path)
    print(f'Saved: {output_path}')


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description='Export reversal results to Excel')
    parser.add_argument('--n', type=int, default=None)
    parser.add_argument('--from-cache', action='store_true', default=True)
    parser.add_argument('--out', type=str, default=None)
    args = parser.parse_args()

    # Output path
    xlsx_path = args.out or os.path.join(OUT_DIR, 'reversal_detailed.xlsx')
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

    # Files
    files = sorted([f for f in os.listdir(AECG_DIR) if f.endswith('.aECG')])
    if args.n is not None:
        files = files[:args.n]

    detector = LimbLeadReversalDetector()
    processor = None if args.from_cache else LimbLeadProcessor(max_samples=MAX_SAMPLES)
    results = []
    patient_names = {}

    print(f'Processing {len(files)} records...')
    t0 = time.time()

    for idx, fname in enumerate(files):
        rec_name = fname.replace('.aECG', '')
        fpath = os.path.join(AECG_DIR, fname)

        # Extract patient name from aECG
        pname = get_patient_name(fpath)
        patient_names[rec_name] = pname

        try:
            if args.from_cache:
                ll_result = load_cached_result(rec_name)
                if ll_result is None:
                    aecg = parse_aecg(fpath, max_samples=MAX_SAMPLES)
                    ll_result, _ = processor.process_record(aecg, record_name=rec_name) if processor else (None, None)
                    if processor is None:
                        processor = LimbLeadProcessor(max_samples=MAX_SAMPLES)
                        aecg = parse_aecg(fpath, max_samples=MAX_SAMPLES)
                        ll_result, _ = processor.process_record(aecg, record_name=rec_name)
            else:
                aecg = parse_aecg(fpath, max_samples=MAX_SAMPLES)
                ll_result, _ = processor.process_record(aecg, record_name=rec_name)

            if ll_result is None:
                continue

            result = detector.detect(ll_result)
            results.append(result)

        except Exception as e:
            print(f'  [{idx+1}] {rec_name}: ERROR {e}')
            continue

        if (idx + 1) % 10 == 0:
            print(f'  [{idx+1}/{len(files)}] ... {time.time()-t0:.0f}s')

    dt = time.time() - t0
    print(f'Completed {len(results)} records in {dt:.0f}s')

    if results:
        export_to_xlsx(results, detector, xlsx_path, patient_names=patient_names)

        # Quick summary
        vc = Counter(r.reversal_type for r in results)
        print(f'\nSummary:')
        for k in ['normal', 'ra_la', 'uncertain']:
            cnt = vc.get(k, 0)
            print(f'  {k:<12}: {cnt:>3} ({cnt/len(results)*100:>5.1f}%)')


if __name__ == '__main__':
    main()
