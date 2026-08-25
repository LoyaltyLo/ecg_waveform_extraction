#!/usr/bin/env python3
"""Batch QRS C1 Processing: HSMM + C1 polarity + QRS plots + XLSX.

For each record:
  1. HSMM segmentation → QRS boundaries
  2. C1 (dominant deflection) polarity per beat
  3. QRS C1 plots (overview + per-beat detail with max/min markers)
  4. XLSX export

Usage:
    python -m ecg_waveform_extraction.src.batch_qrs_c1 --n 50
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

import os, json, time, gc, argparse
from collections import Counter
import numpy as np

from ecg_waveform_extraction.src.limb_lead_processor import (
    LimbLeadProcessor, LIMB_LEADS,
)
from ecg_waveform_extraction.src.utils.aecg_parser import parse_aecg
from ecg_waveform_extraction.src.plot_qrs_c1 import save_qrs_c1_plots
from ecg_waveform_extraction.src.export_reversal_xlsx import get_patient_name

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
AECG_DIR = 'C:/LoyaltyLo/datasets/RA-LA_Reversal/aECG'
OUT_DIR = str(Path(__file__).resolve().parent.parent / 'output/rala_full/_qrs_c1')
MAX_SAMPLES = 4000

HEADER_FONT = Font(name='Consolas', size=10, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_FONT = Font(name='Consolas', size=9)
DATA_ALIGN = Alignment(horizontal='center', vertical='center')
TEXT_ALIGN = Alignment(horizontal='left', vertical='center')
FILL_POS = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
FILL_NEG = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


# ---------------------------------------------------------------------------
# Process one record
# ---------------------------------------------------------------------------
def process_one(fpath, processor):
    """Run HSMM with C1-only, save QRS plots, return summary dict."""
    rec_name = os.path.basename(fpath).replace('.aECG', '')
    rec_dir = os.path.join(OUT_DIR, rec_name)
    os.makedirs(rec_dir, exist_ok=True)

    pname = get_patient_name(fpath)

    aecg = parse_aecg(fpath, max_samples=MAX_SAMPLES)
    ll_result, seg_data = processor.process_record(aecg, record_name=rec_name)

    # Save QRS C1 plots
    save_qrs_c1_plots(seg_data, ll_result, rec_name, rec_dir)

    # Build per-lead summary
    leads_summary = {}
    for ln in LIMB_LEADS:
        lr = ll_result.leads.get(ln)
        if lr is None or lr.n_beats == 0:
            leads_summary[ln] = None
            continue

        pos = lr.polarity_counts.get('positive', 0)
        neg = lr.polarity_counts.get('negative', 0)
        total = max(pos + neg, 1)
        dominant = 'positive' if pos >= neg else 'negative'

        # C1 confidence: average across beats
        confs = [b['confidence'] for b in lr.beats
                if b['polarity'] in ('positive', 'negative')]
        mean_conf = float(np.mean(confs)) if confs else 0.0

        leads_summary[ln] = {
            'n_beats': lr.n_beats,
            'positive': pos, 'negative': neg,
            'pos_pct': round(pos / total * 100, 1),
            'neg_pct': round(neg / total * 100, 1),
            'dominant': dominant,
            'mean_c1_conf': round(mean_conf, 3),
        }

    meas = ll_result.measurements
    return {
        'record': rec_name,
        'patient_name': pname,
        'leads': leads_summary,
        'P_axis': meas.get('P_axis', ''),
        'QRS_axis': meas.get('QRS_axis', ''),
        'HR': meas.get('HR', ''),
        'interpretation': ll_result.interpretation or '',
    }


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
def export_xlsx(summaries, output_path):
    """Generate focused QRS C1 XLSX."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'QRS C1 Polarity'

    # Headers
    headers = [
        'Record', 'Patient Name',
        'I Beats', 'I +', 'I -', 'I +%', 'I -%', 'I Dominant', 'I C1 Conf',
        'II Beats', 'II +', 'II -', 'II +%', 'II -%', 'II Dominant', 'II C1 Conf',
        'III Beats', 'III +', 'III -', 'III Dominant',
        'AVR Beats', 'AVR +', 'AVR -', 'AVR Dominant',
        'AVL Beats', 'AVL +', 'AVL -', 'AVL Dominant',
        'AVF Beats', 'AVF +', 'AVF -', 'AVF Dominant',
        'P_axis', 'QRS_axis', 'HR', 'Interpretation',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for i, s in enumerate(summaries):
        row = i + 2
        values = [s['record'], s['patient_name']]

        for ln in LIMB_LEADS:
            ld = s['leads'].get(ln)
            if ld is None:
                if ln in ('I', 'II'):
                    values += ['', '', '', '', '', '', '']
                else:
                    values += ['', '', '', '']
                continue

            if ln in ('I', 'II'):
                values += [ld['n_beats'], ld['positive'], ld['negative'],
                          ld['pos_pct'], ld['neg_pct'],
                          ld['dominant'], ld['mean_c1_conf']]
            else:
                values += [ld['n_beats'], ld['positive'], ld['negative'],
                          ld['dominant']]

        values += [s['P_axis'], s['QRS_axis'], s['HR'],
                   (s['interpretation'] or '')[:60]]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN if not isinstance(val, str) else TEXT_ALIGN
            cell.border = THIN_BORDER

        # Color Lead I and aVR dominant columns
        for dom_col in [8, 24]:  # I dominant, aVR dominant
            v = ws.cell(row=row, column=dom_col).value
            if v == 'positive':
                ws.cell(row=row, column=dom_col).fill = FILL_POS
            elif v == 'negative':
                ws.cell(row=row, column=dom_col).fill = FILL_NEG

    # Column widths
    widths = [14, 10, 6, 5, 5, 6, 6, 8, 8, 6, 5, 5, 6, 6, 8, 8, 6, 5, 5, 8, 6, 5, 5, 8, 6, 5, 5, 8, 6, 5, 5, 8, 8, 8, 6, 40]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.freeze_panes = 'C2'
    ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(len(headers))}{len(summaries)+1}'

    wb.save(output_path)
    print(f'Saved: {output_path}')


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--n', type=int, default=50)
    parser.add_argument('--start', type=int, default=0)
    args = parser.parse_args()

    files = sorted([f for f in os.listdir(AECG_DIR) if f.endswith('.aECG')])
    if args.start > 0:
        files = files[args.start:]
    files = files[:args.n]
    n_total = len(files)
    os.makedirs(OUT_DIR, exist_ok=True)

    processor = LimbLeadProcessor(max_samples=MAX_SAMPLES)
    summaries = []

    print(f'QRS C1 Processing: {n_total} records')
    t0 = time.time()

    for idx, fname in enumerate(files):
        fpath = os.path.join(AECG_DIR, fname)
        rec_name = fname.replace('.aECG', '')
        print(f'[{idx+1:3d}/{n_total}] {rec_name}...', end=' ', flush=True)
        t1 = time.time()

        try:
            s = process_one(fpath, processor)
            summaries.append(s)

            # Quick stats
            i_ld = s['leads'].get('I') or {}
            avr_ld = s['leads'].get('AVR') or {}
            print(f'OK I:{i_ld.get("dominant","?")} '
                  f'aVR:{avr_ld.get("dominant","?")} '
                  f'({time.time()-t1:.0f}s)')
        except Exception as e:
            print(f'ERROR: {e}')

        gc.collect()

    dt = time.time() - t0
    print(f'\nCompleted {len(summaries)} records in {dt:.0f}s ({dt/n_total:.1f}s/rec)')

    # XLSX
    xlsx_path = os.path.join(OUT_DIR, 'qrs_c1_summary.xlsx')
    export_xlsx(summaries, xlsx_path)

    # Summary
    i_pos = sum(1 for s in summaries if (s['leads'].get('I') or {}).get('dominant') == 'positive')
    i_neg = sum(1 for s in summaries if (s['leads'].get('I') or {}).get('dominant') == 'negative')
    avr_pos = sum(1 for s in summaries if (s['leads'].get('AVR') or {}).get('dominant') == 'positive')
    avr_neg = sum(1 for s in summaries if (s['leads'].get('AVR') or {}).get('dominant') == 'negative')

    print(f'Lead I : +{i_pos}  -{i_neg}')
    print(f'aVR    : +{avr_pos}  -{avr_neg}')
    print(f'Output : {OUT_DIR}/')


if __name__ == '__main__':
    main()
