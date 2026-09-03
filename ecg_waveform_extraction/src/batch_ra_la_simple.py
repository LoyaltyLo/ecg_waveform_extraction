#!/usr/bin/env python3
"""Limb Lead Reversal Detection — RA-LA + LA-LL.

RA-LA Rule (3-step):
  1. Lead I QRS C1: if positive > negative → NORMAL
  2. Lead I QRS negative >= positive: check Lead I T-wave
     - T-wave positive > negative → NORMAL
     - T-wave negative >= positive → proceed to step 3
  3. Check aVR QRS C1:
     - aVR positive > negative → RA-LA REVERSAL

LA-LL Rule (2-step + LAD exclusion):
  1. Lead III QRS negative >= positive + T negative >= positive
  2. P-axis abnormal (not 0-75°) to exclude LAD
  3. aVL↔aVF swap as supporting evidence

RA-LL Rule:
  Lead I, II, III QRS all negative (neg >= pos) + aVR QRS positive (pos > neg)

Usage:
    python -m ecg_waveform_extraction.src.batch_ra_la_simple --n 50
    python -m ecg_waveform_extraction.src.batch_ra_la_simple --n 50 --no-plots
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

import os, json, time, gc, argparse
from collections import Counter
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

from ecg_waveform_extraction.src.limb_lead_processor import LimbLeadProcessor
from ecg_waveform_extraction.src.utils.aecg_parser import parse_aecg
from ecg_waveform_extraction.src.export_reversal_xlsx import get_patient_name
from ecg_waveform_extraction.src.chest_lead_analyzer import (
    ChestLeadAnalyzer, chest_result_to_dict,
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
AECG_DIR = 'C:/LoyaltyLo/datasets/RA-LA_Reversal/aECG'
OUT_DIR = str(Path(__file__).resolve().parent.parent / 'output/rala_full/_ra_la_simple')
MAX_SAMPLES = 16000  # full 10 s at 1 kHz (4000 was the 250 Hz-era value = 4 s)

HEADER_FONT = Font(name='Consolas', size=10, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_FONT = Font(name='Consolas', size=9)
DATA_ALIGN = Alignment(horizontal='center', vertical='center')
TEXT_ALIGN = Alignment(horizontal='left', vertical='center')
FILL_REV = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
FILL_NORM = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
FILL_UNC = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

# ---------------------------------------------------------------------------
# T-wave plotting
# ---------------------------------------------------------------------------
def plot_t_wave(ecg_clean, t_on, t_off, beat_id, lead_name, rec_name,
                save_path, polarity, fs=250.0, dpi=130):
    """Plot single T-wave beat with polarity coloring."""
    margin = int(0.06 * fs)  # 60ms padding
    T = len(ecg_clean)
    ws = max(0, t_on - margin)
    we = min(T - 1, t_off + margin)
    if we - ws < 10:
        return

    t_win = np.arange(ws, we + 1) / fs
    e_win = ecg_clean[ws:we + 1]
    t_t = np.arange(t_on, t_off + 1) / fs
    e_t = ecg_clean[t_on:t_off + 1]

    # T-wave polarity color
    color = '#4caf50' if polarity == 'positive' else '#f44336'
    label = 'T POSITIVE' if polarity == 'positive' else 'T NEGATIVE'

    fig, ax = plt.subplots(figsize=(7, 3))
    ax.plot(t_win, e_win, 'k-', linewidth=0.6, alpha=0.5)
    ax.plot(t_t, e_t, color=color, linewidth=1.2)
    ax.fill_between(t_t, e_t, alpha=0.25, color=color, linewidth=0)

    # Max point
    t_peak_idx = t_on + np.argmax(np.abs(ecg_clean[t_on:t_off + 1]))
    if ws <= t_peak_idx <= we:
        ax.plot(t_peak_idx / fs, ecg_clean[t_peak_idx], 'v',
                color='darkgreen' if polarity == 'positive' else 'darkred',
                markersize=10)

    ax.set_title(f'{rec_name} — Lead {lead_name} Beat {beat_id}  |  T-wave: {label}',
                 fontsize=10, fontweight='bold')
    ax.set_xlabel('Time (s)'); ax.set_ylabel('Amplitude')
    ax.set_xlim(t_win[0], t_win[-1]); ax.grid(True, alpha=0.12)
    fig.tight_layout()
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    fig.savefig(save_path, dpi=dpi, bbox_inches='tight')
    plt.close(fig)


def plot_qrs_c1_beat(ecg_clean, q_on, r_pk, s_off, beat_id, lead_name,
                     rec_name, save_path, polarity, fs=250.0, dpi=130):
    """Plot single QRS beat with max/min markers."""
    margin = int(0.05 * fs)
    T = len(ecg_clean)
    ws = max(0, q_on - margin)
    we = min(T - 1, s_off + margin)
    if we - ws < 10:
        return

    t_win = np.arange(ws, we + 1) / fs
    e_win = ecg_clean[ws:we + 1]
    seg = ecg_clean[q_on:s_off + 1]
    bl_win = int(0.12 * fs)  # 120 ms pre-QRS baseline
    detrend = seg - np.median(ecg_clean[max(0, q_on - bl_win):q_on]) if q_on >= bl_win else seg - np.median(seg[:5])

    pos_max = float(np.max(detrend))
    neg_min = float(np.min(detrend))
    max_idx = q_on + int(np.argmax(detrend))
    min_idx = q_on + int(np.argmin(detrend))

    color = '#4caf50' if polarity == 'positive' else '#f44336'

    fig, ax = plt.subplots(figsize=(7, 3))
    ax.plot(t_win, e_win, 'k-', linewidth=0.6, alpha=0.5)
    t_qrs = np.arange(q_on, s_off + 1) / fs
    ax.plot(t_qrs, ecg_clean[q_on:s_off + 1], color=color, linewidth=1.2)
    ax.fill_between(t_qrs, ecg_clean[q_on:s_off + 1], alpha=0.25, color=color, linewidth=0)

    if ws <= max_idx <= we:
        ax.plot(max_idx / fs, ecg_clean[max_idx], 'rv', markersize=10, zorder=5)
        ax.annotate(f'{pos_max:.2f}', (max_idx / fs, ecg_clean[max_idx]),
                    textcoords='offset points', xytext=(8, 6), fontsize=8, color='darkred')
    if ws <= min_idx <= we:
        ax.plot(min_idx / fs, ecg_clean[min_idx], 'bv', markersize=10, zorder=5)
        ax.annotate(f'{neg_min:.2f}', (min_idx / fs, ecg_clean[min_idx]),
                    textcoords='offset points', xytext=(8, -12), fontsize=8, color='darkblue')

    label = 'POSITIVE' if polarity == 'positive' else 'NEGATIVE'
    ax.text(0.02, 0.95, f'QRS C1: {label}\nMax: {pos_max:.3f}  Min: {neg_min:.3f}',
            transform=ax.transAxes, fontsize=8, va='top', fontfamily='monospace',
            bbox=dict(boxstyle='round,pad=0.4', facecolor='white', alpha=0.8))
    ax.set_title(f'{rec_name} — Lead {lead_name} Beat {beat_id}  |  QRS C1', fontsize=10, fontweight='bold')
    ax.set_xlabel('Time (s)'); ax.set_ylabel('Amplitude')
    ax.set_xlim(t_win[0], t_win[-1]); ax.grid(True, alpha=0.12)
    fig.tight_layout()
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    fig.savefig(save_path, dpi=dpi, bbox_inches='tight')
    plt.close(fig)


# ---------------------------------------------------------------------------
def process_record(fpath, processor, save_plots=True):
    """Run HSMM, apply RA-LA + LA-LL detection, save plots."""
    rec_name = os.path.basename(fpath).replace('.aECG', '')
    pname = get_patient_name(fpath)

    aecg = parse_aecg(fpath, max_samples=MAX_SAMPLES)
    ll_result, seg_data = processor.process_record(aecg, record_name=rec_name)

    # ---- Gather data ----
    lr_I = ll_result.leads.get('I')
    lr_II = ll_result.leads.get('II')
    lr_III = ll_result.leads.get('III')
    lr_avr = ll_result.leads.get('AVR')
    lr_avl = ll_result.leads.get('AVL')
    lr_avf = ll_result.leads.get('AVF')
    if lr_I is None or lr_I.n_beats == 0:
        return None
    if lr_avr is None or lr_avr.n_beats == 0:
        return None

    # Lead I
    i_qrs_pos = lr_I.polarity_counts.get('positive', 0)
    i_qrs_neg = lr_I.polarity_counts.get('negative', 0)
    i_t_pos = lr_I.t_polarity_counts.get('positive', 0)
    i_t_neg = lr_I.t_polarity_counts.get('negative', 0)

    # Lead II
    ii_qrs_pos = lr_II.polarity_counts.get('positive', 0) if lr_II else 0
    ii_qrs_neg = lr_II.polarity_counts.get('negative', 0) if lr_II else 0

    # aVR
    avr_pos = lr_avr.polarity_counts.get('positive', 0)
    avr_neg = lr_avr.polarity_counts.get('negative', 0)

    # Lead III
    iii_qrs_pos = lr_III.polarity_counts.get('positive', 0) if lr_III else 0
    iii_qrs_neg = lr_III.polarity_counts.get('negative', 0) if lr_III else 0
    iii_t_pos = lr_III.t_polarity_counts.get('positive', 0) if lr_III else 0
    iii_t_neg = lr_III.t_polarity_counts.get('negative', 0) if lr_III else 0

    # aVL / aVF (for swap detection)
    avl_qrs_pos = lr_avl.polarity_counts.get('positive', 0) if lr_avl else 0
    avl_qrs_neg = lr_avl.polarity_counts.get('negative', 0) if lr_avl else 0
    avf_qrs_pos = lr_avf.polarity_counts.get('positive', 0) if lr_avf else 0
    avf_qrs_neg = lr_avf.polarity_counts.get('negative', 0) if lr_avf else 0

    # aVL↔aVF swap signal
    avl_avf_swapped = (avl_qrs_pos > avl_qrs_neg and avf_qrs_neg > avf_qrs_pos)

    # ---- RA-LA Detection ----
    ra_la = 'normal'
    ra_la_reason = ''
    if i_qrs_pos > i_qrs_neg:
        ra_la = 'normal'
        ra_la_reason = f'I QRS +{i_qrs_pos} > -{i_qrs_neg}'
    elif i_t_pos > i_t_neg:
        ra_la = 'normal'
        ra_la_reason = f'I QRS -{i_qrs_neg} >= +{i_qrs_pos}, but T +{i_t_pos} > -{i_t_neg}'
    elif i_t_neg >= i_t_pos:
        if avr_pos > avr_neg:
            ra_la = 'reversed'
            ra_la_reason = f'I QRS(-{i_qrs_neg}/+{i_qrs_pos}) T(-{i_t_neg}/+{i_t_pos}) both down, aVR +{avr_pos}>-{avr_neg}'
        else:
            ra_la = 'uncertain'
            ra_la_reason = f'I QRS/T both down, aVR +{avr_pos}/-{avr_neg} not majority pos'
    else:
        ra_la = 'uncertain'
        ra_la_reason = f'I QRS -{i_qrs_neg} >= +{i_qrs_pos}, T tie (+{i_t_pos} vs -{i_t_neg})'

    # ---- LA-LL Detection ----
    la_ll = 'normal'
    la_ll_reason = ''
    # P-axis for LAD exclusion (LAD has normal P-axis, true LA-LL has abnormal P-axis)
    p_axis = ll_result.measurements.get('P_axis')
    p_axis_normal = (p_axis is not None and 0 <= p_axis <= 75)
    qrs_axis = ll_result.measurements.get('QRS_axis')

    if lr_III is None:
        la_ll = 'N/A'
        la_ll_reason = 'No Lead III data'
    else:
        # Core: Lead III QRS + T both inverted
        iii_both_inverted = (iii_qrs_neg >= iii_qrs_pos and iii_t_neg >= iii_t_pos
                             and (iii_qrs_neg > 0 or iii_t_neg > 0))
        # Confirmation: aVL↔aVF swap
        swap_confirm = avl_avf_swapped
        # P-axis must be abnormal to exclude LAD (LAD has normal P-axis)
        not_lad = not p_axis_normal

        if iii_both_inverted and not_lad:
            # True LA-LL: III inverted + abnormal P-axis (excludes LAD)
            extra = ''
            if swap_confirm:
                extra = f' + aVL↔aVF swap'
            if qrs_axis is not None:
                extra += f' QRS_axis={qrs_axis}'
            la_ll = 'reversed'
            la_ll_reason = (f'III QRS(-{iii_qrs_neg}/+{iii_qrs_pos}) T(-{iii_t_neg}/+{iii_t_pos}) '
                            f'P_axis={p_axis} (abnormal→not LAD){extra}')
        elif iii_both_inverted and p_axis_normal:
            # III inverted but P-axis normal → likely LAD, not true LA-LL
            la_ll = 'normal'
            la_ll_reason = (f'III QRS(-{iii_qrs_neg}/+{iii_qrs_pos}) T(-{iii_t_neg}/+{iii_t_pos}) '
                            f'but P_axis={p_axis} normal → likely LAD, not LA-LL')
        elif swap_confirm and not p_axis_normal:
            la_ll = 'uncertain'
            la_ll_reason = (f'aVL↔aVF swap + P_axis={p_axis} abnormal, '
                            f'but III QRS/T not both inverted')
        elif iii_qrs_pos > iii_qrs_neg:
            la_ll = 'normal'
            la_ll_reason = f'III QRS +{iii_qrs_pos} > -{iii_qrs_neg}'
        else:
            la_ll = 'normal'
            la_ll_reason = f'III QRS -{iii_qrs_neg}/+{iii_qrs_pos}, T +{iii_t_pos}/-{iii_t_neg}'

    # ---- RA-LL Detection ----
    ra_ll = 'normal'
    ra_ll_reason = ''
    if ii_qrs_neg >= ii_qrs_pos and iii_qrs_neg >= iii_qrs_pos and i_qrs_neg >= i_qrs_pos \
            and (i_qrs_neg > 0 or ii_qrs_neg > 0 or iii_qrs_neg > 0) \
            and avr_pos > avr_neg:
        ra_ll = 'reversed'
        ra_ll_reason = (f'I(-{i_qrs_neg}/+{i_qrs_pos}) II(-{ii_qrs_neg}/+{ii_qrs_pos}) '
                        f'III(-{iii_qrs_neg}/+{iii_qrs_pos}) all neg, aVR(+{avr_pos}/-{avr_neg}) pos')
    elif i_qrs_pos > i_qrs_neg or ii_qrs_pos > ii_qrs_neg:
        ra_ll = 'normal'
        ra_ll_reason = 'Not all limb leads inverted'
    else:
        ra_ll = 'normal'
        ra_ll_reason = f'I(-{i_qrs_neg}) II(-{ii_qrs_neg}) III(-{iii_qrs_neg}) aVR(+{avr_pos}/-{avr_neg})'

    # ---- Combined verdict ----
    if ra_la == 'reversed':
        verdict = 'RA-LA'
    elif ra_ll == 'reversed':
        verdict = 'RA-LL'
    elif la_ll == 'reversed':
        verdict = 'LA-LL'
    elif ra_la == 'uncertain' or (i_qrs_neg >= i_qrs_pos and i_t_neg >= i_t_pos):
        verdict = 'uncertain'
    else:
        verdict = 'normal'

    # ---- Save plots ----
    if save_plots:
        rec_dir = os.path.join(OUT_DIR, rec_name)

        def _save_lead_plots(lead_name, lr, sd):
            if sd is None or lr is None: return
            ecg = sd['filtered_ecg']
            fs = sd.get('fs', 250.0)  # threaded from _process_lead
            ld = os.path.join(rec_dir, f'lead_{lead_name}')
            qd = os.path.join(ld, 'qrs_beats')
            td = os.path.join(ld, 't_beats')
            os.makedirs(qd, exist_ok=True); os.makedirs(td, exist_ok=True)
            for b in lr.beats:
                bid = b['beat_id']
                if b['q_onset'] > 0 and b['s_offset'] > b['q_onset']:
                    plot_qrs_c1_beat(ecg, b['q_onset'], b['r_peak'], b['s_offset'],
                                     bid, lead_name, rec_name,
                                     os.path.join(qd, f'beat_{bid:03d}.png'), b['polarity'],
                                     fs=fs)
            for tw in lr.t_waves:
                bid = tw['beat_id']
                if tw['onset'] > 0 and tw['offset'] > tw['onset']:
                    plot_t_wave(ecg, tw['onset'], tw['offset'], bid, lead_name,
                                rec_name, os.path.join(td, f'beat_{bid:03d}.png'), tw['polarity'],
                                fs=fs)

        _save_lead_plots('I', lr_I, seg_data.get('I'))
        _save_lead_plots('II', lr_II, seg_data.get('II'))
        _save_lead_plots('III', lr_III, seg_data.get('III'))

    # ---- Chest lead analysis ----
    chest_analyzer = ChestLeadAnalyzer(fs=aecg.get('fs', 250.0), max_samples=MAX_SAMPLES)
    chest = chest_analyzer.analyze(aecg)
    chest_dict = chest_result_to_dict(chest)

    meas = ll_result.measurements
    result_dict = {
        'record': rec_name,
        'patient_name': pname,
        'verdict': verdict,
        'RA_LA': ra_la, 'RA_LA_reason': ra_la_reason,
        'RA_LL': ra_ll, 'RA_LL_reason': ra_ll_reason,
        'LA_LL': la_ll, 'LA_LL_reason': la_ll_reason,
        'I_QRS_pos': i_qrs_pos, 'I_QRS_neg': i_qrs_neg,
        'II_QRS_pos': ii_qrs_pos, 'II_QRS_neg': ii_qrs_neg,
        'I_T_pos': i_t_pos, 'I_T_neg': i_t_neg,
        'avr_pos': avr_pos, 'avr_neg': avr_neg,
        'III_QRS_pos': iii_qrs_pos, 'III_QRS_neg': iii_qrs_neg,
        'III_T_pos': iii_t_pos, 'III_T_neg': iii_t_neg,
        'aVL_QRS_pos': avl_qrs_pos, 'aVL_QRS_neg': avl_qrs_neg,
        'aVF_QRS_pos': avf_qrs_pos, 'aVF_QRS_neg': avf_qrs_neg,
        'aVL_aVF_swap': avl_avf_swapped,
        'I_beats': lr_I.n_beats, 'II_beats': lr_II.n_beats if lr_II else 0,
        'avr_beats': lr_avr.n_beats, 'III_beats': lr_III.n_beats if lr_III else 0,
        'P_axis': meas.get('P_axis', ''),
        'QRS_axis': meas.get('QRS_axis', ''),
        'HR': meas.get('HR', ''),
        'interpretation': ll_result.interpretation or '',
        # Chest lead data
        'chest': chest_dict,
    }
    return result_dict


# ---------------------------------------------------------------------------
def export_xlsx(results, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'RA-LA + LA-LL'

    headers = [
        'Record', 'Patient Name', 'Verdict',
        'RA-LA', 'RA-LA Reason',
        'RA-LL', 'RA-LL Reason',
        'LA-LL', 'LA-LL Reason',
        'I QRS +', 'I QRS -', 'II QRS +', 'II QRS -',
        'I T +', 'I T -',
        'aVR +', 'aVR -',
        'III QRS +', 'III QRS -',
        'III T +', 'III T -',
        'aVL QRS +', 'aVL QRS -',
        'aVF QRS +', 'aVF QRS -',
        'aVL↔aVF',
        'I Beats', 'II Beats', 'aVR Beats', 'III Beats',
        'V1 R', 'V1 S', 'V2 R', 'V2 S', 'V3 R', 'V3 S',
        'V4 R', 'V4 S', 'V5 R', 'V5 S', 'V6 R', 'V6 S',
        'V1 R/S', 'V2 R/S', 'V3 R/S', 'V4 R/S', 'V5 R/S', 'V6 R/S',
        'V1-V2 SWAP', 'R-Progression', 'Transition', 'Chest Flags',
        'P_axis', 'QRS_axis', 'HR', 'Interpretation',
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN; c.border = THIN_BORDER

    for i, r in enumerate(results):
        row = i + 2
        vals = [
            r['record'], r['patient_name'], r['verdict'],
            r['RA_LA'], r['RA_LA_reason'],
            r['RA_LL'], r['RA_LL_reason'],
            r['LA_LL'], r['LA_LL_reason'],
            r['I_QRS_pos'], r['I_QRS_neg'],
            r['II_QRS_pos'], r['II_QRS_neg'],
            r['I_T_pos'], r['I_T_neg'],
            r['avr_pos'], r['avr_neg'],
            r['III_QRS_pos'], r['III_QRS_neg'],
            r['III_T_pos'], r['III_T_neg'],
            r['aVL_QRS_pos'], r['aVL_QRS_neg'],
            r['aVF_QRS_pos'], r['aVF_QRS_neg'],
            'YES' if r['aVL_aVF_swap'] else '',
            r['I_beats'], r['II_beats'], r['avr_beats'], r['III_beats'],
            r['chest']['V1_R'], r['chest']['V1_S'],
            r['chest']['V2_R'], r['chest']['V2_S'],
            r['chest']['V3_R'], r['chest']['V3_S'],
            r['chest']['V4_R'], r['chest']['V4_S'],
            r['chest']['V5_R'], r['chest']['V5_S'],
            r['chest']['V6_R'], r['chest']['V6_S'],
            r['chest']['V1_R/S'], r['chest']['V2_R/S'],
            r['chest']['V3_R/S'], r['chest']['V4_R/S'],
            r['chest']['V5_R/S'], r['chest']['V6_R/S'],
            'YES' if r['chest']['V1-V2_SWAP'] else '',
            r['chest']['R_Progression'],
            r['chest']['Transition'],
            r['chest']['Chest_Flags'],
            r['P_axis'], r['QRS_axis'], r['HR'],
            (r['interpretation'] or '')[:80],
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = DATA_FONT
            c.alignment = DATA_ALIGN if not isinstance(v, str) else TEXT_ALIGN
            c.border = THIN_BORDER

        vc = ws.cell(row=row, column=3)
        if r['verdict'] in ('RA-LA', 'RA-LL', 'LA-LL'):
            vc.fill = FILL_REV
            vc.font = Font(name='Consolas', size=9, bold=True, color='B71C1C')
        elif r['verdict'] == 'normal':
            vc.fill = FILL_NORM
        else:
            vc.fill = FILL_UNC

    widths = [14, 10, 10, 8, 40, 8, 40, 8, 40, 6, 6, 6, 6, 6, 6, 6, 6, 7, 7, 6, 6, 7, 7, 6, 6, 6, 6, 6, 6, 6,
              5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5,
              5, 5, 5, 5, 5, 5,
              8, 8, 10, 30, 8, 9, 6, 40]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(len(headers))}{len(results)+1}'

    wb.save(path)
    print(f'Saved: {path}')


# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--n', type=int, default=None)
    parser.add_argument('--out', type=str, default=None)
    parser.add_argument('--no-plots', action='store_true')
    args = parser.parse_args()

    global OUT_DIR
    if args.out:
        OUT_DIR = os.path.dirname(args.out) if os.path.dirname(args.out) else OUT_DIR
    os.makedirs(OUT_DIR, exist_ok=True)

    files = sorted([f for f in os.listdir(AECG_DIR) if f.endswith('.aECG')])
    if args.n:
        files = files[:args.n]

    xlsx_path = args.out or os.path.join(OUT_DIR, 'ra_la_qrs_t_wave.xlsx')

    processor = LimbLeadProcessor(max_samples=MAX_SAMPLES)
    results = []
    save_plots = not args.no_plots

    print(f'RA-LA QRS+T Wave Rule: {len(files)} records')
    if save_plots:
        print(f'Plots: QRS + T-wave beat details will be saved')
    t0 = time.time()

    for idx, fname in enumerate(files):
        fpath = os.path.join(AECG_DIR, fname)
        rec = fname.replace('.aECG', '')
        print(f'[{idx+1:3d}/{len(files)}] {rec}...', end=' ', flush=True)
        t1 = time.time()

        try:
            r = process_record(fpath, processor, save_plots=save_plots)
            if r:
                results.append(r)
                print(f'{r["verdict"]:<10} QRS:+{r["I_QRS_pos"]}/-{r["I_QRS_neg"]} '
                      f'T:+{r["I_T_pos"]}/-{r["I_T_neg"]} '
                      f'aVR:+{r["avr_pos"]}/-{r["avr_neg"]} '
                      f'({time.time()-t1:.0f}s)')
            else:
                print('SKIP')
        except Exception as e:
            print(f'ERROR: {e}')
        gc.collect()

    dt = time.time() - t0
    print(f'\nCompleted {len(results)} records in {dt:.0f}s ({dt/len(results):.1f}s/rec)')

    export_xlsx(results, xlsx_path)

    vc = Counter(r['verdict'] for r in results)
    ra_la_vc = Counter(r['RA_LA'] for r in results)
    ra_ll_vc = Counter(r['RA_LL'] for r in results)
    la_ll_vc = Counter(r['LA_LL'] for r in results)
    print(f'\nResults:')
    for k in ['normal', 'RA-LA', 'RA-LL', 'LA-LL', 'uncertain']:
        cnt = vc.get(k, 0)
        print(f'  {k:<12}: {cnt:>4} ({cnt/len(results)*100:>5.1f}%)')
    print(f'\n  RA-LA detail:')
    for k in ['normal', 'reversed', 'uncertain']:
        print(f'    {k:<12}: {ra_la_vc.get(k, 0):>4}')
    print(f'  RA-LL detail:')
    for k in ['normal', 'reversed']:
        print(f'    {k:<12}: {ra_ll_vc.get(k, 0):>4}')
    print(f'  LA-LL detail:')
    for k in ['normal', 'reversed', 'N/A']:
        print(f'    {k:<12}: {la_ll_vc.get(k, 0):>4}')

    step2_saved = [r for r in results if 'but T' in r['RA_LA_reason']]
    print(f'\n  RA-LA T-wave saved: {len(step2_saved)}')
    print(f'  RA-LL confirmed: {ra_ll_vc.get("reversed", 0)}')
    print(f'  LA-LL confirmed: {la_ll_vc.get("reversed", 0)}')


if __name__ == '__main__':
    main()
