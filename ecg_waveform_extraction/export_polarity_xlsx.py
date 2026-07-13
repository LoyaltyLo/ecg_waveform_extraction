"""Generate comprehensive polarity detection Excel report.

Sheets:
  1. Summary        — 总体统计、各方法汇总
  2. Per-Record     — 每条记录的详细检测结果
  3. P-Wave Method  — P波极性检测细节
  4. Lead Compare   — 导联对比检测细节
  5. P-Axis Method  — P波电轴检测细节
  6. Combined       — 三方法综合投票结果
  7. Conflicting    — 各方法结果冲突的记录
"""

import sys
sys.path.insert(0, 'c:/LoyaltyLo/PythonProjects/ECG_engineering')

import os, json, re
import numpy as np
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.series import DataPoint

# ---- Config ----
AECG_DIR = 'C:/LoyaltyLo/datasets/RA-LA_Reversal/aECG'
OUT_DIR = 'c:/LoyaltyLo/PythonProjects/ECG_engineering/ecg_waveform_extraction/output_rala_full'
XLSX_PATH = 'c:/LoyaltyLo/PythonProjects/ECG_engineering/ecg_waveform_extraction/output_rala_full/polarity_detection_report.xlsx'

# ---- Styles ----
HEADER_FONT = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

NORMAL_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
REVERSED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
UNCERTAIN_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
BORDERLINE_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
TITLE_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
SECTION_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
SECTION_FONT = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=13)

THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

DATA_FONT = Font(name='Consolas', size=10)
DATA_ALIGN = Alignment(horizontal='center', vertical='center')
TEXT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
NUM_ALIGN = Alignment(horizontal='right', vertical='center')


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, align=DATA_ALIGN):
    cell = ws.cell(row=row, column=col)
    cell.font = DATA_FONT
    cell.alignment = align
    cell.border = THIN_BORDER


def auto_width(ws, min_w=8, max_w=50):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_w
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split('\n')
                max_line = max(len(line) for line in lines)
                max_len = max(max_len, min(max_line + 2, max_w))
        ws.column_dimensions[col_letter].width = max_len


# =====================================================================
# Collect all data
# =====================================================================
print("Collecting polarity data from all records...")

records = sorted([
    d for d in os.listdir(OUT_DIR)
    if os.path.isdir(os.path.join(OUT_DIR, d)) and d[:1].isdigit()
])

all_data = []

for idx, rec in enumerate(records):
    aecg_path = os.path.join(AECG_DIR, rec + '.aECG')

    entry = {
        'idx': idx + 1,
        'record': rec,
        'file_exists': os.path.exists(aecg_path),
        'p_axis': None,
        'p_axis_polarity': None,
        'lead_I_inverted': None,
        'lead_II_P_inverted': None,
        'lead_comparison_polarity': None,
        'pw_polarity': None,
        'pw_n_positive': 0,
        'pw_n_negative': 0,
        'pw_n_total': 0,
        'pw_mean_area': None,
        'ecg_interpretation': '',
        'hr_bpm': None,
        'pr_ms': None,
        'qrs_ms': None,
        'qt_ms': None,
        'consensus': None,
        'confidence': 0.0,
    }

    if not os.path.exists(aecg_path):
        all_data.append(entry)
        continue

    # Parse aECG
    for enc in ['utf-8', 'gbk', 'latin-1']:
        try:
            with open(aecg_path, 'r', encoding=enc) as f:
                content = f.read()
            if '<?xml' in content[:100]:
                break
        except:
            continue

    # -- P-axis --
    m = re.search(r'MDC_ECG_ANGLE_P_FRONT.*?<value[^>]*value="([^"]+)"', content, re.DOTALL)
    p_axis = float(m.group(1)) if m else None
    entry['p_axis'] = p_axis
    if p_axis is not None:
        if 0 <= p_axis <= 75:
            entry['p_axis_polarity'] = 'normal'
        elif p_axis > 100 or p_axis < -30:
            entry['p_axis_polarity'] = 'reversed'
        elif 75 < p_axis <= 100:
            entry['p_axis_polarity'] = 'borderline_right'
        else:
            entry['p_axis_polarity'] = 'atypical'

    # -- Lead I / II comparison --
    digits = list(re.finditer(r'<digits[^>]*>([^<]*)</digits>', content))
    if len(digits) >= 2:
        sig_I = np.array([float(x) for x in digits[0].group(1).split()])[:4000]
        sig_II = np.array([float(x) for x in digits[1].group(1).split()])[:4000]

        li_absmax = np.argmax(np.abs(sig_I - np.median(sig_I)))
        entry['lead_I_inverted'] = (sig_I[li_absmax] - np.median(sig_I)) < 0

        early_II = sig_II[200:600]
        entry['lead_II_P_inverted'] = bool(np.sum(early_II - np.median(early_II)) < 0)

        li_inv = entry['lead_I_inverted']
        lii_p = entry['lead_II_P_inverted']
        if li_inv and lii_p:
            entry['lead_comparison_polarity'] = 'reversed'
        elif not li_inv and not lii_p:
            entry['lead_comparison_polarity'] = 'normal'
        elif li_inv:
            entry['lead_comparison_polarity'] = 'likely_reversed'
        else:
            entry['lead_comparison_polarity'] = 'uncertain'

    # -- Global measurements --
    for key, pat in {
        'hr_bpm': r'MDC_ECG_HEART_RATE.*?<value[^>]*value="([^"]+)"[^>]*unit="bpm"',
        'pr_ms': r'MDC_ECG_TIME_PD_PR.*?<value[^>]*value="([^"]+)"[^>]*unit="ms"',
        'qrs_ms': r'MDC_ECG_TIME_PD_QRS\b(?!c).*?<value[^>]*value="([^"]+)"[^>]*unit="ms"',
        'qt_ms': r'MDC_ECG_TIME_PD_QT\b(?!c).*?<value[^>]*value="([^"]+)"[^>]*unit="ms"',
    }.items():
        m = re.search(pat, content, re.DOTALL)
        if m:
            entry[key] = float(m.group(1))

    # -- Interpretation --
    interp = re.search(
        r'MDC_ECG_INTERPRETATION_STATEMENT.*?xsi:type="ST"[^>]*>([^<]+)</value>',
        content, re.DOTALL)
    if interp:
        entry['ecg_interpretation'] = interp.group(1).strip().replace('\n', '; ')

    # -- P-wave polarity from HSMM --
    pw_path = os.path.join(OUT_DIR, rec, 'p_waves.json')
    ecg_path = os.path.join(OUT_DIR, rec, 'filtered_ecg.npy')
    if os.path.exists(pw_path) and os.path.exists(ecg_path):
        with open(pw_path) as f:
            pws = json.load(f)
        ecg = np.load(ecg_path)
        p_areas = []
        for pw in pws:
            onset, offset = pw['onset_sample'], pw['offset_sample']
            if onset < 0 or offset <= onset:
                continue
            seg = ecg[onset:offset + 1]
            bl = np.mean(ecg[max(0, onset - 50):onset]) if onset >= 50 else np.mean(seg[:10])
            p_areas.append(np.sum(seg - bl))
        if p_areas:
            entry['pw_n_total'] = len(p_areas)
            entry['pw_n_positive'] = sum(1 for a in p_areas if a > 0)
            entry['pw_n_negative'] = sum(1 for a in p_areas if a < 0)
            entry['pw_mean_area'] = round(float(np.mean(p_areas)), 4)
            entry['pw_polarity'] = 'normal' if np.mean(p_areas) > 0 else 'reversed'

    # -- Consensus voting --
    votes = {'reversed': 0, 'normal': 0}
    if entry['p_axis_polarity'] == 'normal':
        votes['normal'] += 1
    elif entry['p_axis_polarity'] in ('reversed',):
        votes['reversed'] += 1
    if entry['lead_comparison_polarity'] == 'normal':
        votes['normal'] += 1
    elif entry['lead_comparison_polarity'] in ('reversed', 'likely_reversed'):
        votes['reversed'] += 1
    if entry['pw_polarity'] == 'normal':
        votes['normal'] += 1
    elif entry['pw_polarity'] == 'reversed':
        votes['reversed'] += 1

    total_votes = votes['reversed'] + votes['normal']
    if total_votes >= 2:
        entry['consensus'] = 'reversed' if votes['reversed'] > votes['normal'] else 'normal'
        entry['confidence'] = round(max(votes['reversed'], votes['normal']) / total_votes, 2)
    elif total_votes == 1:
        entry['consensus'] = 'reversed' if votes['reversed'] > votes['normal'] else 'normal'
        entry['confidence'] = 0.5
    else:
        entry['consensus'] = 'uncertain'
        entry['confidence'] = 0.0

    all_data.append(entry)

    if (idx + 1) % 150 == 0:
        print(f"  {idx+1}/{len(records)}...", flush=True)

print(f"  Total: {len(all_data)} records collected")


# =====================================================================
# Create workbook
# =====================================================================
print("\nBuilding Excel report...")
wb = Workbook()

# ==================== Sheet 1: Summary ====================
ws1 = wb.active
ws1.title = 'Summary'
ws1.sheet_properties.tabColor = '2F5496'

# Title
ws1.merge_cells('A1:H1')
title_cell = ws1['A1']
title_cell.value = 'RA-LA Reversal aECG Dataset — Polarity Detection Report'
title_cell.font = Font(name='Microsoft YaHei', bold=True, size=16, color='1F4E79')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 35

ws1.merge_cells('A2:H2')
ws1['A2'].value = f'Total Records: {len(all_data)} | Generated from HSMM segmentation + aECG annotations'
ws1['A2'].font = Font(name='Microsoft YaHei', size=10, color='666666', italic=True)
ws1['A2'].alignment = Alignment(horizontal='center')
ws1.row_dimensions[2].height = 22

# --- Section: Dataset Overview ---
row = 4
ws1.merge_cells(f'A{row}:H{row}')
ws1.cell(row=row, column=1, value='📊 数据集概览').font = SECTION_FONT
ws1.cell(row=row, column=1).fill = SECTION_FILL
for c in range(1, 9):
    ws1.cell(row=row, column=c).fill = SECTION_FILL
ws1.row_dimensions[row].height = 28

row = 5
overview = [
    ['总记录数', len(all_data), '', '采样率', '1000 Hz', '', '导联数', '12'],
    ['HSMM 检测心拍总数', sum(1 for d in all_data for _ in [1] if d.get('pw_n_total', 0) > 0), '', '分析时长/条', '4 秒', '', '数据类型', 'HL7 aECG XML'],
    ['P 波电轴来源', 'aECG 系统自动测量', '', 'P 波极性来源', 'HSMM 分割 + Lead II 积分', '', 'QRS 来源', 'Lead I/II 原始信号'],
]
for i, ov in enumerate(overview):
    for j, val in enumerate(ov):
        c = ws1.cell(row=row + i, column=j + 1, value=val)
        c.font = Font(name='Microsoft YaHei', size=10)
        c.alignment = Alignment(horizontal='center' if j % 2 == 0 else 'left')
        c.border = THIN_BORDER
        if j % 3 == 0:
            c.fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
            c.font = Font(name='Microsoft YaHei', bold=True, size=10)

# --- Section: Polarity Methods ---
row = 10
ws1.merge_cells(f'A{row}:H{row}')
ws1.cell(row=row, column=1, value='🔍 三种检测方法').font = SECTION_FONT
for c in range(1, 9):
    ws1.cell(row=row, column=c).fill = SECTION_FILL
ws1.row_dimensions[row].height = 28

row = 11
method_headers = ['方法', '原理', '正常指标', '反接指标', '数据来源', '可靠性', '检出反接数', '反接占比']
for j, h in enumerate(method_headers):
    c = ws1.cell(row=row, column=j + 1, value=h)
style_header_row(ws1, row, len(method_headers))
ws1.row_dimensions[row].height = 25

n_p_axis_rev = sum(1 for d in all_data if d.get('p_axis_polarity') == 'reversed')
n_pw_rev = sum(1 for d in all_data if d.get('pw_polarity') == 'reversed')
n_lead_rev = sum(1 for d in all_data if d.get('lead_comparison_polarity') in ('reversed', 'likely_reversed'))
total = max(len(all_data), 1)

methods_data = [
    ['① P 波电轴', '额面 P 波向量角度', '0° ~ 75°', '>100° 或 <-30°',
     'aECG XML (MDC_ECG_ANGLE_P_FRONT)', '⭐⭐⭐ 金标准',
     n_p_axis_rev, f'{n_p_axis_rev/total*100:.1f}%'],
    ['② P 波极性 (HSMM)', 'Lead II P 波净面积符号', '正向 (↑)', '负向 (↓)',
     'HSMM 分割 + filtered_ecg.npy', '⭐⭐⭐ 需窦性心律',
     n_pw_rev, f'{n_pw_rev/total*100:.1f}%'],
    ['③ Lead I QRS 方向', 'Lead I QRS 主波方向', '正向', '负向',
     '原始 Lead I 信号', '⭐⭐ 有假阳性',
     n_lead_rev, f'{n_lead_rev/total*100:.1f}%'],
]
for i, md in enumerate(methods_data):
    for j, val in enumerate(md):
        c = ws1.cell(row=row + 1 + i, column=j + 1, value=val)
        c.font = Font(name='Microsoft YaHei', size=10)
        c.alignment = Alignment(horizontal='center' if j != 1 else 'left', vertical='center', wrap_text=True)
        c.border = THIN_BORDER
    ws1.row_dimensions[row + 1 + i].height = 40

# --- Section: Consensus Results ---
row = 17
ws1.merge_cells(f'A{row}:H{row}')
ws1.cell(row=row, column=1, value='📋 综合投票结果').font = SECTION_FONT
for c in range(1, 9):
    ws1.cell(row=row, column=c).fill = SECTION_FILL
ws1.row_dimensions[row].height = 28

row = 18
consensus_headers = ['判定结果', '记录数', '占比', '说明']
for j, h in enumerate(consensus_headers):
    ws1.cell(row=row, column=j + 1, value=h)
style_header_row(ws1, row, len(consensus_headers))

n_normal = sum(1 for d in all_data if d.get('consensus') == 'normal')
n_reversed = sum(1 for d in all_data if d.get('consensus') == 'reversed')
n_uncertain = sum(1 for d in all_data if d.get('consensus') == 'uncertain')

consensus_data = [
    ['极性正常 (Normal)', n_normal, f'{n_normal/total*100:.1f}%',
     '三方法中 ≥2 票投"正常"'],
    ['极性反接 (Reversed)', n_reversed, f'{n_reversed/total*100:.1f}%',
     'RA-LA 电极左右手反接，需重新采集'],
    ['不确定 (Uncertain)', n_uncertain, f'{n_uncertain/total*100:.1f}%',
     '方法结果矛盾或仅有 1 票，需人工复核'],
]

fills = [NORMAL_FILL, REVERSED_FILL, UNCERTAIN_FILL]
for i, cd in enumerate(consensus_data):
    for j, val in enumerate(cd):
        c = ws1.cell(row=row + 1 + i, column=j + 1, value=val)
        c.font = Font(name='Microsoft YaHei', size=11, bold=(j == 0))
        c.alignment = Alignment(horizontal='center' if j < 3 else 'left')
        c.border = THIN_BORDER
        c.fill = fills[i]

# --- Section: Pie Chart ---
row = 24
ws1.merge_cells(f'A{row}:C{row}')
ws1.cell(row=row, column=1, value='综合判定分布').font = Font(name='Microsoft YaHei', bold=True, size=11)

# Data for chart
ws1.cell(row=row + 1, column=1, value='分类')
ws1.cell(row=row + 1, column=2, value='数量')
ws1.cell(row=row + 2, column=1, value='极性正常')
ws1.cell(row=row + 2, column=2, value=n_normal)
ws1.cell(row=row + 3, column=1, value='极性反接')
ws1.cell(row=row + 3, column=2, value=n_reversed)
ws1.cell(row=row + 4, column=1, value='不确定')
ws1.cell(row=row + 4, column=2, value=n_uncertain)

pie = PieChart()
pie.title = 'RA-LA 极性综合判定'
pie.style = 10
labels = Reference(ws1, min_col=1, min_row=row + 2, max_row=row + 4)
data = Reference(ws1, min_col=2, min_row=row + 1, max_row=row + 4)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.width = 18
pie.height = 12
# Colors
colors = ['C6EFCE', 'FFC7CE', 'FFEB9C']
for i, color in enumerate(colors):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = color
    pie.series[0].data_points.append(pt)

ws1.add_chart(pie, f'E{row}')
ws1.row_dimensions[row].height = 280

# --- Section: Key Statistics ---
row = 26
ws1.merge_cells(f'A{row}:H{row}')
ws1.cell(row=row, column=1, value='📈 关键统计').font = SECTION_FONT
for c in range(1, 9):
    ws1.cell(row=row, column=c).fill = SECTION_FILL

row = 27
valid_axes = [d['p_axis'] for d in all_data if d.get('p_axis') is not None]
normal_axes = [a for a in valid_axes if 0 <= a <= 75]
reversed_axes = [a for a in valid_axes if a > 100 or a < -30]
hr_vals = [d['hr_bpm'] for d in all_data if d.get('hr_bpm') is not None]
pr_vals = [d['pr_ms'] for d in all_data if d.get('pr_ms') is not None]
qrs_vals = [d['qrs_ms'] for d in all_data if d.get('qrs_ms') is not None]

stats = [
    ['P 波电轴均值', f'{np.mean(valid_axes):.1f}°', '正常组均值', f'{np.mean(normal_axes):.1f}°', '反接组均值', f'{np.mean(reversed_axes):.1f}°'],
    ['正常组范围', f'{np.min(normal_axes):.0f}° ~ {np.max(normal_axes):.0f}°', '反接组范围', f'{np.min(reversed_axes):.0f}° ~ {np.max(reversed_axes):.0f}°', '', ''],
    ['HR 均值', f'{np.mean(hr_vals):.1f} bpm', 'PR 均值', f'{np.mean(pr_vals):.1f} ms', 'QRS 均值', f'{np.mean(qrs_vals):.1f} ms'],
]
stat_headers = ['指标', '数值', '指标', '数值', '指标', '数值']
for j, h in enumerate(stat_headers):
    ws1.cell(row=row, column=j + 1, value=h)
style_header_row(ws1, row, 6)

for i, st in enumerate(stats):
    for j, val in enumerate(st):
        c = ws1.cell(row=row + 1 + i, column=j + 1, value=val)
        c.font = Font(name='Consolas', size=10)
        c.alignment = Alignment(horizontal='center')
        c.border = THIN_BORDER
        if j % 2 == 0:
            c.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

auto_width(ws1, min_w=12, max_w=55)


# ==================== Sheet 2: Per-Record ====================
ws2 = wb.create_sheet('Per-Record')
ws2.sheet_properties.tabColor = '4472C4'

headers2 = [
    '#', '记录名', '综合判定', '置信度', 'P波电轴(°)', 'P轴判定',
    'P波极性(HSMM)', 'P波+数', 'P波-数', 'P波净面积',
    'Lead I反相', 'Lead II P波反相', '导联对比判定',
    'HR(bpm)', 'PR(ms)', 'QRS(ms)', 'QT(ms)', 'ECG解读'
]
for j, h in enumerate(headers2):
    ws2.cell(row=1, column=j + 1, value=h)
style_header_row(ws2, 1, len(headers2))
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f'A1:{get_column_letter(len(headers2))}{len(all_data) + 1}'

for i, d in enumerate(all_data):
    r = i + 2
    vals = [
        d['idx'], d['record'], d.get('consensus', ''), d.get('confidence', 0),
        d.get('p_axis'), d.get('p_axis_polarity', ''),
        d.get('pw_polarity', ''), d.get('pw_n_positive', 0), d.get('pw_n_negative', 0),
        d.get('pw_mean_area'),
        'Yes' if d.get('lead_I_inverted') else ('No' if d.get('lead_I_inverted') is False else ''),
        'Yes' if d.get('lead_II_P_inverted') else ('No' if d.get('lead_II_P_inverted') is False else ''),
        d.get('lead_comparison_polarity', ''),
        d.get('hr_bpm'), d.get('pr_ms'), d.get('qrs_ms'), d.get('qt_ms'),
        d.get('ecg_interpretation', '')[:200],
    ]
    for j, val in enumerate(vals):
        c = ws2.cell(row=r, column=j + 1, value=val)
        c.font = DATA_FONT
        c.alignment = TEXT_ALIGN if j in (len(vals) - 1,) else DATA_ALIGN
        c.border = THIN_BORDER

    # Color rows by consensus
    consensus = d.get('consensus', '')
    if consensus == 'reversed':
        for j in range(1, 4):
            ws2.cell(row=r, column=j).fill = REVERSED_FILL
    elif consensus == 'normal':
        for j in range(1, 4):
            ws2.cell(row=r, column=j).fill = NORMAL_FILL
    elif consensus == 'uncertain':
        for j in range(1, 4):
            ws2.cell(row=r, column=j).fill = UNCERTAIN_FILL

    # Color P-axis and P-wave columns
    if d.get('p_axis_polarity') == 'reversed':
        ws2.cell(row=r, column=5).fill = REVERSED_FILL
        ws2.cell(row=r, column=6).fill = REVERSED_FILL
    if d.get('pw_polarity') == 'reversed':
        ws2.cell(row=r, column=7).fill = REVERSED_FILL

auto_width(ws2, min_w=8, max_w=40)
ws2.column_dimensions[get_column_letter(len(headers2))].width = 55  # ECG interpretation column wider


# ==================== Sheet 3: P-Wave Method Detail ====================
ws3 = wb.create_sheet('P-Wave Method')
ws3.sheet_properties.tabColor = '70AD47'

headers3 = ['记录名', 'P波数', '正向(+个)', '负向(-个)', '净面积均值', 'P波极性', '综合判定', '一致?']
for j, h in enumerate(headers3):
    ws3.cell(row=1, column=j + 1, value=h)
style_header_row(ws3, 1, len(headers3))
ws3.freeze_panes = 'A2'

for i, d in enumerate(all_data):
    r = i + 2
    consensus = d.get('consensus', '')
    pw_pol = d.get('pw_polarity', '')
    agree = '✓' if (pw_pol == consensus and consensus in ('normal', 'reversed')) else ('—' if not pw_pol else '✗')

    vals = [
        d['record'], d.get('pw_n_total', 0), d.get('pw_n_positive', 0),
        d.get('pw_n_negative', 0), d.get('pw_mean_area'),
        pw_pol, consensus, agree,
    ]
    for j, val in enumerate(vals):
        c = ws3.cell(row=r, column=j + 1, value=val)
        c.font = DATA_FONT
        c.alignment = DATA_ALIGN
        c.border = THIN_BORDER

    if pw_pol == 'reversed':
        for j in range(1, 7):
            ws3.cell(row=r, column=j).fill = REVERSED_FILL
    elif pw_pol == 'normal' and d.get('pw_n_total', 0) > 0:
        for j in range(1, 7):
            ws3.cell(row=r, column=j).fill = NORMAL_FILL

    if agree == '✗':
        ws3.cell(row=r, column=8).fill = REVERSED_FILL
        ws3.cell(row=r, column=8).font = Font(name='Consolas', bold=True, size=12, color='C00000')
    elif agree == '✓':
        ws3.cell(row=r, column=8).font = Font(name='Consolas', bold=True, size=12, color='006100')

auto_width(ws3)


# ==================== Sheet 4: Lead Comparison Detail ====================
ws4 = wb.create_sheet('Lead Comparison')
ws4.sheet_properties.tabColor = 'ED7D31'

headers4 = ['记录名', 'Lead I 反相', 'Lead II P反相', 'Lead I 净标志', '导联对比判定', '综合判定', '一致?']
for j, h in enumerate(headers4):
    ws4.cell(row=1, column=j + 1, value=h)
style_header_row(ws4, 1, len(headers4))
ws4.freeze_panes = 'A2'

for i, d in enumerate(all_data):
    r = i + 2
    consensus = d.get('consensus', '')
    lc_pol = d.get('lead_comparison_polarity', '')
    agree = '✓' if (lc_pol == consensus or
                    (lc_pol == 'likely_reversed' and consensus == 'reversed') or
                    (lc_pol == 'normal' and consensus == 'normal')) else ('—' if not lc_pol else '✗')

    vals = [
        d['record'],
        'Yes (反相)' if d.get('lead_I_inverted') else ('No' if d.get('lead_I_inverted') is False else 'N/A'),
        'Yes' if d.get('lead_II_P_inverted') else ('No' if d.get('lead_II_P_inverted') is False else 'N/A'),
        'Inverted' if d.get('lead_I_inverted') else ('Normal' if d.get('lead_I_inverted') is False else ''),
        lc_pol, consensus, agree,
    ]
    for j, val in enumerate(vals):
        c = ws4.cell(row=r, column=j + 1, value=val)
        c.font = DATA_FONT
        c.alignment = DATA_ALIGN
        c.border = THIN_BORDER
    if agree == '✗':
        ws4.cell(row=r, column=7).fill = REVERSED_FILL

auto_width(ws4)


# ==================== Sheet 5: P-Axis Method Detail ====================
ws5 = wb.create_sheet('P-Axis Method')
ws5.sheet_properties.tabColor = '5B9BD5'

headers5 = ['记录名', 'P波电轴(°)', '判定类别', '综合判定', '一致?', 'HR(bpm)', '备注']
for j, h in enumerate(headers5):
    ws5.cell(row=1, column=j + 1, value=h)
style_header_row(ws5, 1, len(headers5))
ws5.freeze_panes = 'A2'

for i, d in enumerate(all_data):
    r = i + 2
    consensus = d.get('consensus', '')
    pa_pol = d.get('p_axis_polarity', '')
    agree = '✓' if (pa_pol == consensus or (pa_pol == 'borderline_right' and consensus in ('normal', 'uncertain'))) else ('—' if not pa_pol else '✗')

    # Axis category
    axis = d.get('p_axis')
    if axis is not None:
        if 0 <= axis <= 75:
            axis_cat = '正常 (0-75°)'
        elif axis > 100:
            axis_cat = '反接 (>100°)'
        elif axis < -30:
            axis_cat = '反接 (<-30°)'
        elif 75 < axis <= 100:
            axis_cat = '临界右偏 (75-100°)'
        else:
            axis_cat = f'不典型 ({axis:.0f}°)'
    else:
        axis_cat = 'N/A'

    vals = [
        d['record'], axis, axis_cat,
        consensus, agree, d.get('hr_bpm'),
        'P轴右偏→RA-LA反接' if (axis is not None and axis > 100) else '',
    ]
    for j, val in enumerate(vals):
        c = ws5.cell(row=r, column=j + 1, value=val)
        c.font = DATA_FONT
        c.alignment = DATA_ALIGN if j != 6 else TEXT_ALIGN
        c.border = THIN_BORDER
    if agree == '✗':
        ws5.cell(row=r, column=5).fill = REVERSED_FILL
    if d.get('p_axis_polarity') == 'reversed':
        ws5.cell(row=r, column=2).fill = REVERSED_FILL
        ws5.cell(row=r, column=3).fill = REVERSED_FILL

auto_width(ws5)


# ==================== Sheet 6: Combined Voting ====================
ws6 = wb.create_sheet('Combined Voting')
ws6.sheet_properties.tabColor = '7030A0'

headers6 = [
    '记录名', 'P轴投票', '导联对比投票', 'P波极性投票',
    '正常票', '反接票', '总票数', '综合判定', '置信度',
    'HR(bpm)', 'ECG解读'
]
for j, h in enumerate(headers6):
    ws6.cell(row=1, column=j + 1, value=h)
style_header_row(ws6, 1, len(headers6))
ws6.freeze_panes = 'A2'

for i, d in enumerate(all_data):
    r = i + 2
    # Count votes
    v_normal = sum([
        1 if d.get('p_axis_polarity') == 'normal' else 0,
        1 if d.get('lead_comparison_polarity') == 'normal' else 0,
        1 if d.get('pw_polarity') == 'normal' else 0,
    ])
    v_reversed = sum([
        1 if d.get('p_axis_polarity') == 'reversed' else 0,
        1 if d.get('lead_comparison_polarity') in ('reversed', 'likely_reversed') else 0,
        1 if d.get('pw_polarity') == 'reversed' else 0,
    ])

    p_axis_vote = '正常' if d.get('p_axis_polarity') == 'normal' else ('反接' if d.get('p_axis_polarity') == 'reversed' else '—')
    lead_vote = '正常' if d.get('lead_comparison_polarity') == 'normal' else ('反接' if d.get('lead_comparison_polarity') in ('reversed', 'likely_reversed') else '—')
    pw_vote = '正常' if d.get('pw_polarity') == 'normal' else ('反接' if d.get('pw_polarity') == 'reversed' else '—')

    vals = [
        d['record'], p_axis_vote, lead_vote, pw_vote,
        v_normal, v_reversed, v_normal + v_reversed,
        d.get('consensus', ''), d.get('confidence', 0),
        d.get('hr_bpm'), d.get('ecg_interpretation', '')[:180],
    ]
    for j, val in enumerate(vals):
        c = ws6.cell(row=r, column=j + 1, value=val)
        c.font = DATA_FONT
        c.alignment = DATA_ALIGN if j < len(vals) - 1 else TEXT_ALIGN
        c.border = THIN_BORDER

    # Highlight rows
    consensus = d.get('consensus', '')
    if consensus == 'reversed':
        for j in [8]:
            ws6.cell(row=r, column=j).fill = REVERSED_FILL
    elif consensus == 'normal':
        for j in [8]:
            ws6.cell(row=r, column=j).fill = NORMAL_FILL

    # Highlight misaligned votes
    if v_reversed > 0 and consensus == 'normal':
        ws6.cell(row=r, column=6).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    if v_normal > 0 and consensus == 'reversed':
        ws6.cell(row=r, column=5).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    # Data bar for confidence
    ws6.cell(row=r, column=9).number_format = '0%'

auto_width(ws6)
ws6.column_dimensions[get_column_letter(len(headers6))].width = 50


# ==================== Sheet 7: Conflicting Records ====================
ws7 = wb.create_sheet('Conflicting')
ws7.sheet_properties.tabColor = 'FF0000'

headers7 = [
    '记录名', '综合判定', '冲突类型', 'P轴判定', '导联对比', 'P波极性',
    'HR', 'PR', 'QRS', 'ECG解读（摘要）'
]
for j, h in enumerate(headers7):
    ws7.cell(row=1, column=j + 1, value=h)
style_header_row(ws7, 1, len(headers7))
ws7.freeze_panes = 'A2'

conflict_row = 2
for d in all_data:
    pa = d.get('p_axis_polarity', '')
    lc = d.get('lead_comparison_polarity', '')
    pw = d.get('pw_polarity', '')

    # Detect conflicts
    methods = []
    if pa == 'reversed': methods.append('P-axis=反接')
    elif pa == 'normal': methods.append('P-axis=正常')
    if lc in ('reversed', 'likely_reversed'): methods.append('Lead=反接')
    elif lc == 'normal': methods.append('Lead=正常')
    if pw == 'reversed': methods.append('P-wave=反接')
    elif pw == 'normal': methods.append('P-wave=正常')

    # Conflict = mixed votes among methods
    has_normal = any('正常' in m for m in methods)
    has_reversed = any('反接' in m for m in methods)

    if not has_normal or not has_reversed:
        continue  # No conflict

    conflict_type = ' & '.join(methods)

    vals = [
        d['record'], d.get('consensus', '?'), conflict_type,
        pa, lc, pw,
        d.get('hr_bpm'), d.get('pr_ms'), d.get('qrs_ms'),
        (d.get('ecg_interpretation', '') or '')[:150],
    ]
    for j, val in enumerate(vals):
        c = ws7.cell(row=conflict_row, column=j + 1, value=val)
        c.font = DATA_FONT
        c.alignment = DATA_ALIGN if j < len(vals) - 1 else TEXT_ALIGN
        c.border = THIN_BORDER
        c.fill = UNCERTAIN_FILL
    conflict_row += 1

if conflict_row == 2:
    ws7.cell(row=2, column=1, value='无冲突记录 — 所有方法结果一致')
    ws7.merge_cells('A2:J2')
    ws7['A2'].font = Font(name='Microsoft YaHei', bold=True, size=12, color='006100')
    ws7['A2'].fill = NORMAL_FILL

auto_width(ws7)
ws7.column_dimensions[get_column_letter(len(headers7))].width = 55


# =====================================================================
# Freeze and save
# =====================================================================
for ws in [ws2, ws3, ws4, ws5, ws6, ws7]:
    ws.freeze_panes = ws.cell(row=2, column=1)

print(f"\nSaving to: {XLSX_PATH}")
wb.save(XLSX_PATH)

import os
size_mb = os.path.getsize(XLSX_PATH) / (1024 * 1024)
print(f"Done! File size: {size_mb:.1f} MB")

# Print summary for terminal
print(f"\n{'='*55}")
print(f"  POLARITY DETECTION REPORT")
print(f"{'='*55}")
print(f"  Sheets: Summary, Per-Record, P-Wave Method, Lead Comparison,")
print(f"          P-Axis Method, Combined Voting, Conflicting")
print(f"  Total records: {len(all_data)}")
print(f"  Normal: {n_normal} ({n_normal/total*100:.1f}%)")
print(f"  Reversed: {n_reversed} ({n_reversed/total*100:.1f}%)")
print(f"  Uncertain: {n_uncertain} ({n_uncertain/total*100:.1f}%)")
if conflict_row > 2:
    print(f"  Conflicting methods: {conflict_row - 2} records")
print(f"  Report: {XLSX_PATH}")
print(f"{'='*55}")
